"""Generate the ad-hoc / benchmark Excel parity oracle for the Rust port.

Runs the reference Python `AdHocExcelWriter.write_research` with a FIXED,
deterministic benchmark input (peer comparison, comparative → summary stats),
reads the produced `.xlsx` back with openpyxl, and dumps a cell-level snapshot
(`value` / `formula` / `fill`, mirroring `characterize_xlsx`) to
`tieout/excel_snapshots/ADHOC_bench_snapshot.json`.

The only non-deterministic cell — the footer `Generated: <timestamp> | ...`
line — is normalized to a pinned string so the gate is reproducible; the Rust
gate builds with the same pinned string.

Run:  py tieout/build_adhoc_oracle.py
"""
import json
import os
import sys
from pathlib import Path

REPO = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(REPO))

SNAP_DIR = REPO / "tieout/excel_snapshots"
XLSX_DIR = REPO / "tests/snapshots"
XLSX_DIR.mkdir(parents=True, exist_ok=True)
SNAP_DIR.mkdir(parents=True, exist_ok=True)

# Pinned "generated" stamp — the Rust gate uses the identical string.
PINNED_GENERATED = "Generated: 2026-01-01 00:00 | Source: SEC EDGAR / yfinance / Company filings"


def characterize_xlsx(path: Path) -> dict:
    """Mirror of tieout/build_full_is_oracle.py::characterize_xlsx."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            cells = []
            for cell in row:
                c = {"ref": cell.coordinate}
                if cell.value is not None:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        c["formula"] = val
                    else:
                        # Normalize the non-deterministic footer timestamp.
                        if isinstance(val, str) and val.startswith("Generated: "):
                            val = PINNED_GENERATED
                        c["value"] = val
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                    c["fill"] = fill.fgColor.rgb
                if cell.hyperlink:
                    c["hyperlink"] = cell.hyperlink.target
                cells.append(c)
            if cells:
                rows.append({"row": row[0].row, "cells": cells})
        sheets[name] = rows
    wb.close()
    return sheets


def fixed_benchmark():
    """Deterministic peer-benchmark input. Columns grouped so the group banner
    exercises both multi-column merges and profitability ratios. Values are
    round so f64 parity is exact. Mirrored verbatim on the Rust side."""
    from src.research.output_writer import ColumnSpec

    columns = [
        ColumnSpec("ticker", "Ticker", "text", width=10, is_label=True),
        ColumnSpec("revenue", "Revenue", "dollar", width=14, units="USD millions",
                   group="Financials", definition="Total net revenue, latest FY"),
        ColumnSpec("ebitda", "EBITDA", "dollar", width=14, units="USD millions",
                   group="Financials"),
        ColumnSpec("net_income", "Net Income", "dollar", width=14, units="USD millions",
                   group="Financials"),
        ColumnSpec("ebitda_margin", "EBITDA Margin", "percent", width=13,
                   group="Profitability"),
        ColumnSpec("net_margin", "Net Margin", "percent", width=13,
                   group="Profitability"),
        ColumnSpec("ev_ebitda", "EV / EBITDA", "multiple", width=12,
                   group="Valuation"),
    ]

    rows = [
        {"ticker": "AAPL", "revenue": 391035.0, "ebitda": 134661.0, "net_income": 93736.0,
         "ebitda_margin": 0.3444, "net_margin": 0.2397, "ev_ebitda": 22.5},
        {"ticker": "MSFT", "revenue": 245122.0, "ebitda": 133558.0, "net_income": 88136.0,
         "ebitda_margin": 0.5449, "net_margin": 0.3596, "ev_ebitda": 24.1},
        {"ticker": "GOOGL", "revenue": 350018.0, "ebitda": 123456.0, "net_income": 100118.0,
         "ebitda_margin": 0.3527, "net_margin": 0.2860, "ev_ebitda": 15.3},
        {"ticker": "AMZN", "revenue": 637959.0, "ebitda": 111500.0, "net_income": 59248.0,
         "ebitda_margin": 0.1748, "net_margin": 0.0929, "ev_ebitda": 18.7},
        {"ticker": "META", "revenue": 164501.0, "ebitda": 87200.0, "net_income": 62360.0,
         "ebitda_margin": 0.5301, "net_margin": 0.3791, "ev_ebitda": 14.2},
    ]

    sources = {
        ("AAPL", "revenue"): "AAPL 10-K FY2024 p.31 (us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax)",
        ("MSFT", "revenue"): "MSFT 10-K FY2024 p.55",
        ("AMZN", "net_income"): "AMZN 10-K FY2024 p.38",
    }
    return columns, rows, sources


def main() -> None:
    from src.research.output_writer import AdHocExcelWriter

    columns, rows, sources = fixed_benchmark()
    xlsx_path = XLSX_DIR / "ADHOC_bench.xlsx"
    writer = AdHocExcelWriter(output_dir=str(XLSX_DIR))
    writer.write_research(
        title="Big Tech - Peer Benchmark (FY2024)",
        rows=rows,
        columns=columns,
        grain="company",
        is_comparative=True,
        needs_sort_filter=True,
        sources=sources,
        filename="ADHOC_bench.xlsx",
    )

    sheets = characterize_xlsx(xlsx_path)
    dst = SNAP_DIR / "ADHOC_bench_snapshot.json"
    dst.write_text(json.dumps({"sheets": sheets}, indent=2), encoding="utf-8")
    n_cells = sum(len(c["cells"]) for rs in sheets.values() for c in rs)
    print(f"  ADHOC bench: sheets={list(sheets.keys())} rows_with_cells="
          f"{sum(len(rs) for rs in sheets.values())} cells={n_cells} -> {dst.name}")


if __name__ == "__main__":
    main()
