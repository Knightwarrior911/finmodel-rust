"""Generate valuation-tab parity oracle (WACC / DCF / Sensitivities / Cover).

Uses the committed SAND_ST empty-IS snapshot's model_output as historicals,
builds a full IS structure, computes WACC (synthetic 2-peer set) + DCF with
FIXED market inputs, writes the workbook, and characterizes it.

Also embeds the serialized WACCOutput / DCFOutput / peer source so the Rust
gate can rehydrate the exact valuation payload and only test the writer.

Output: tieout/excel_snapshots/SAND_ST_val_full_snapshot.json
Run:    py tieout/build_val_oracle.py
"""
from __future__ import annotations

import json
import os
import sys
from dataclasses import asdict
from pathlib import Path

os.environ["FINMODEL_DEV_MOCK"] = "1"
REPO = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(REPO))

SNAP_DIR = REPO / "tieout/excel_snapshots"
XLSX_DIR = REPO / "tests/snapshots"
XLSX_DIR.mkdir(parents=True, exist_ok=True)

SAFE = "SAND_ST"
META = {
    "ticker": "SAND.ST",
    "name": "Sandvik AB",
    "ccy": "SEK",
    "fye": "Dec",
}

# Frozen market inputs — must match the Rust gate constants.
RF = 0.04
SHARE_PX = 220.0
OWN_BETA = 1.15


def characterize_xlsx(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1
        ):
            cells = []
            for cell in row:
                c = {"ref": cell.coordinate}
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        c["formula"] = cell.value
                    else:
                        c["value"] = cell.value
                fill = cell.fill
                if (
                    fill
                    and fill.fgColor
                    and fill.fgColor.rgb
                    and fill.fgColor.rgb != "00000000"
                ):
                    c["fill"] = fill.fgColor.rgb
                if cell.hyperlink:
                    c["hyperlink"] = cell.hyperlink.target
                cells.append(c)
            if cells:
                rows.append({"row": row[0].row, "cells": cells})
        sheets[name] = rows
    wb.close()
    return sheets


def hist_slice(stmt: dict, nh: int) -> dict:
    return {k: (v[:nh] if isinstance(v, list) else v) for k, v in stmt.items()}


def main() -> None:
    from schemas.financial_data import (
        Peer,
        PeerSet,
        PublicCompPeer,
        PublicCompsOutput,
        CompMultipleStats,
        ReconciledFinancialData,
        ModelConfig,
    )
    from src.engine import ModelEngine
    from src.writer import ExcelWriter
    from src.verifier import verify
    from src.assumptions import build_assumptions_block
    from src.is_builder import build_is_structure
    from src.utils import compute_historical_periods
    from src.wacc import compute_wacc
    from src.dcf import compute_dcf

    snap = json.loads((SNAP_DIR / f"{SAFE}_snapshot.json").read_text(encoding="utf-8"))
    mo_s = snap["model_output"]
    periods = snap["periods"]
    nh = sum(1 for p in periods if p.endswith("A"))
    hist_periods = [p for p in periods if p.endswith("A")]

    reconciled = ReconciledFinancialData(
        ticker=META["ticker"],
        company_name=META["name"],
        currency=META["ccy"],
        fiscal_year_end=META["fye"],
        periods=hist_periods,
        income_statement=hist_slice(mo_s["income_statement"], nh),
        balance_sheet=hist_slice(mo_s["balance_sheet"], nh),
        cash_flow_statement=hist_slice(mo_s["cash_flow_statement"], nh),
        notes={},
        sources={},
        flags=[],
    )
    cfg = ModelConfig(
        ticker=META["ticker"],
        company_name=META["name"],
        domicile="non-US",
        currency=META["ccy"],
        fiscal_year_end=META["fye"],
        periods_historical=nh,
        periods_projected=5,
        sector="standard",
    )
    derive_engine = ModelEngine(reconciled, cfg)
    hist = derive_engine._derive_assumptions()
    hp = compute_historical_periods(cfg.fiscal_year_end, cfg.periods_historical)
    ly = int(hp[-1][:4])
    pp = [f"{ly + i + 1}E" for i in range(cfg.periods_projected)]

    class _Stub:
        assumptions = hist
        periods = hp + pp

    assumptions = build_assumptions_block(
        _Stub(), cfg.ticker, sector=cfg.sector, reconciled=reconciled
    )
    # Freeze market inputs so the oracle is deterministic (no yfinance).
    assumptions.risk_free_rate = RF
    assumptions.current_share_price = SHARE_PX
    assumptions.mid_year_convention = True

    engine = ModelEngine(reconciled, cfg, assumptions_block=assumptions)
    mo = engine.build()
    report = verify(mo, sector=cfg.sector)

    _is = mo.income_statement
    has_cogs = any(v and v != 0 for v in (_is.get("cogs") or []))
    has_rd = any(v and v != 0 for v in (_is.get("rd") or []))
    has_sga = any(v and v != 0 for v in (_is.get("sga") or []))
    is_structure = build_is_structure(
        "standard", has_cogs=has_cogs, has_rd=has_rd, has_sga=has_sga
    )

    # Synthetic 2-peer set (deterministic).
    peers = [
        Peer(
            ticker="PEER1",
            name="Peer One",
            market_cap=50_000.0,
            enterprise_value=55_000.0,
            levered_beta=1.20,
            de_ratio=0.35,
            tax_rate=0.21,
            rationale="synthetic",
        ),
        Peer(
            ticker="PEER2",
            name="Peer Two",
            market_cap=40_000.0,
            enterprise_value=48_000.0,
            levered_beta=1.05,
            de_ratio=0.40,
            tax_rate=0.22,
            rationale="synthetic",
        ),
    ]
    shares = assumptions.shares_diluted or 0.0
    if not shares:
        sh = (_is.get("shares_diluted") or [0.0])[-1] or 0.0
        shares = sh
    mkt_cap = SHARE_PX * shares
    debt = (mo.balance_sheet.get("long_term_debt") or [0.0])[-1] or 0.0
    tax = (
        assumptions.base.tax_rate_pct[0]
        if assumptions.base.tax_rate_pct
        else 0.21
    )
    peer_set = PeerSet(
        target_ticker=META["ticker"],
        target_market_cap=mkt_cap,
        target_de_ratio=assumptions.target_de_ratio,
        peers=peers,
        source="synthetic",
    )
    wacc_output = compute_wacc(
        peer_set,
        target_market_cap=mkt_cap,
        target_debt=debt,
        risk_free_rate=RF,
        equity_risk_premium=assumptions.equity_risk_premium,
        cost_of_debt_pretax=assumptions.cost_of_debt_pretax,
        target_tax_rate=tax,
        target_de_ratio=assumptions.target_de_ratio,
        fallback_beta=OWN_BETA,
        sector="standard",
    )
    dcf_output = compute_dcf(
        mo, META["ticker"], wacc_output, assumptions, tv_method=1
    )


    # Synthetic public comps payload (deterministic) for Comps Peers / Summary tabs.
    peers_pc = [
        PublicCompPeer(
            ticker="PEER1", name="Peer One", country="SE", currency="SEK", tier=1,
            share_price=100.0, shares_diluted=500.0, market_cap=50_000.0,
            total_debt=5_000.0, cash=2_000.0, enterprise_value=53_000.0,
            week52_high=120.0, week52_low=80.0,
            ltm_revenue=12_000.0, ltm_ebitda=3_000.0, ltm_ebit=2_400.0,
            ltm_net_income=1_800.0, ltm_eps_diluted=3.6,
            ntm_revenue=12_600.0, ntm_ebitda=3_200.0,
            fy1_revenue=12_400.0, fy1_ebitda=3_100.0,
            fy2_revenue=13_000.0, fy2_ebitda=3_400.0,
            ntm_eps=3.8, fy1_eps=3.7,
            ev_rev_ltm=4.4, ev_ebitda_ltm=17.7, ev_ebit_ltm=22.1, pe_ltm=27.8,
            ev_rev_ntm=4.2, ev_ebitda_ntm=16.6, ev_rev_fy1=4.3, ev_ebitda_fy1=17.1,
            ev_rev_fy2=4.1, ev_ebitda_fy2=15.6, pe_ntm=26.3, pe_fy1=27.0,
            rationale="synthetic",
        ),
        PublicCompPeer(
            ticker="PEER2", name="Peer Two", country="SE", currency="SEK", tier=1,
            share_price=80.0, shares_diluted=500.0, market_cap=40_000.0,
            total_debt=8_000.0, cash=1_500.0, enterprise_value=46_500.0,
            week52_high=100.0, week52_low=60.0,
            ltm_revenue=10_000.0, ltm_ebitda=2_500.0, ltm_ebit=2_000.0,
            ltm_net_income=1_400.0, ltm_eps_diluted=2.8,
            ntm_revenue=10_500.0, ntm_ebitda=2_700.0,
            fy1_revenue=10_300.0, fy1_ebitda=2_600.0,
            fy2_revenue=11_000.0, fy2_ebitda=2_900.0,
            ntm_eps=3.0, fy1_eps=2.9,
            ev_rev_ltm=4.7, ev_ebitda_ltm=18.6, ev_ebit_ltm=23.3, pe_ltm=28.6,
            ev_rev_ntm=4.4, ev_ebitda_ntm=17.2, ev_rev_fy1=4.5, ev_ebitda_fy1=17.9,
            ev_rev_fy2=4.2, ev_ebitda_fy2=16.0, pe_ntm=26.7, pe_fy1=27.6,
            rationale="synthetic",
        ),
    ]
    def _stats(name, vals):
        vs = sorted(vals)
        n = len(vs)
        def pct(p):
            if n == 1: return vs[0]
            k = (n - 1) * p
            f = int(k); c = min(f + 1, n - 1)
            return vs[f] + (vs[c] - vs[f]) * (k - f)
        return CompMultipleStats(
            multiple_name=name, values=vs,
            min=vs[0], p25=pct(0.25), median=pct(0.5), mean=sum(vs)/n,
            p75=pct(0.75), max=vs[-1], count=n,
        )
    stats = {
        "ev_rev_ltm": _stats("EV/Revenue (LTM)", [4.4, 4.7]),
        "ev_ebitda_ltm": _stats("EV/EBITDA (LTM)", [17.7, 18.6]),
        "ev_ebit_ltm": _stats("EV/EBIT (LTM)", [22.1, 23.3]),
        "pe_ltm": _stats("P/E (LTM)", [27.8, 28.6]),
        "ev_rev_ntm": _stats("EV/Revenue (NTM)", [4.2, 4.4]),
        "ev_ebitda_ntm": _stats("EV/EBITDA (NTM)", [16.6, 17.2]),
        "pe_ntm": _stats("P/E (NTM)", [26.3, 26.7]),
    }
    # Implied prices from median EV/EBITDA on target ebitda.
    t_ebitda = (mo.income_statement.get("ebitda") or [0,0])[-1] or 2500.0
    t_debt = debt
    t_cash = (mo.balance_sheet.get("cash") or [0])[-1] or 0.0
    t_sh = shares or 100.0
    med = stats["ev_ebitda_ltm"].median
    p25 = stats["ev_ebitda_ltm"].p25
    p75 = stats["ev_ebitda_ltm"].p75
    def _px(mult):
        ev = t_ebitda * mult
        eq = ev - (t_debt - t_cash)
        return round(eq / t_sh, 2) if t_sh else 0.0
    public_comps = PublicCompsOutput(
        target_ticker=META["ticker"],
        target_company_name=META["name"],
        as_of_date="2025-01-15",
        target_revenue=(mo.income_statement.get("revenue") or [0])[-1] or 0.0,
        target_ebitda=t_ebitda,
        target_ebit=(mo.income_statement.get("ebit") or [0])[-1] or 0.0,
        target_net_income=(mo.income_statement.get("net_income") or [0])[-1] or 0.0,
        target_total_debt=t_debt,
        target_cash=t_cash,
        target_shares_diluted=t_sh,
        peers=peers_pc,
        excluded=[("SKIP1", "too small")],
        stats=stats,
        implied_price_low=_px(p25),
        implied_price_median=_px(med),
        implied_price_high=_px(p75),
        source="synthetic",
    )

    xlsx = XLSX_DIR / f"{SAFE}_val_full.xlsx"
    ExcelWriter(
        mo,
        report,
        META["name"],
        str(xlsx),
        sources={},
        currency=META["ccy"],
        dcf=dcf_output,
        comps=None,
        assumptions=assumptions,
        ticker=META["ticker"],
        fiscal_year_end=META["fye"],
        wacc=wacc_output,
        peer_set=peer_set,
        public_comps=public_comps,
        sector="standard",
        is_structure=is_structure,
    ).write()

    sheets = characterize_xlsx(xlsx)
    out = {
        "ticker": META["ticker"],
        "company": META["name"],
        "currency": META["ccy"],
        "periods": mo.periods,
        "sheets": sheets,
        "model_output": {
            "periods": mo.periods,
            "income_statement": mo.income_statement,
            "balance_sheet": mo.balance_sheet,
            "cash_flow_statement": mo.cash_flow_statement,
        },
        "wacc_output": asdict(wacc_output),
        "dcf_output": asdict(dcf_output),
        "peer_source": peer_set.source,
        "public_comps": asdict(public_comps),
        "market": {
            "risk_free_rate": RF,
            "current_share_price": SHARE_PX,
            "target_de_ratio": assumptions.target_de_ratio,
            "equity_risk_premium": assumptions.equity_risk_premium,
            "cost_of_debt_pretax": assumptions.cost_of_debt_pretax,
            "shares_diluted": shares,
            "tax_rate": tax,
            "mid_year_convention": True,
        },
        "is_flags": {"has_cogs": has_cogs, "has_rd": has_rd, "has_sga": has_sga},
    }
    dst = SNAP_DIR / f"{SAFE}_val_full_snapshot.json"
    dst.write_text(
        json.dumps(out, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
    )
    print(
        f"  {SAFE} [val]: sheets={list(sheets)} "
        f"DCF rows={len(sheets.get('DCF', []))} "
        f"WACC rows={len(sheets.get('WACC', []))} -> {dst.name}"
    )


if __name__ == "__main__":
    main()
