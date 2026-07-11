"""Format-parity measurement: compare per-cell visual finish (font bold/italic/
color) of the Rust-rendered workbook against the writer.py oracle.

The snapshot/value gates are blind to formatting; this script is the format
oracle (openpyxl-side, like the other Python oracles). It reports the % of
cells whose bold/italic/font-color match, per statement sheet.

Prereqs (both written to tests/snapshots/):
  * SAND_ST_full.xlsx  — `py tieout/build_full_is_oracle.py`
  * SAND_ST_rust.xlsx  — `cargo test -p fm-excel --test render_dump`

Run: py tieout/diff_formats.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).parent.parent
PY_XLSX = REPO / "tests/snapshots/SAND_ST_full.xlsx"
RUST_XLSX = REPO / "tests/snapshots/SAND_ST_rust.xlsx"
SHEETS = None  # None → auto-detect every sheet present in both workbooks

def norm_color(c) -> str | None:
    rgb = getattr(c, "rgb", None) if c else None
    return rgb[-6:].upper() if isinstance(rgb, str) else None


def load(path: Path) -> dict:
    wb = openpyxl.load_workbook(path)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                cells[c.coordinate] = {
                    "bold": bool(c.font.bold),
                    "italic": bool(c.font.italic),
                    "color": norm_color(c.font.color),
                }
        out[name] = cells
    return out


def main() -> int:
    if not PY_XLSX.exists() or not RUST_XLSX.exists():
        print("Missing inputs. Generate both xlsx first (see module docstring).")
        return 2
    py, rs = load(PY_XLSX), load(RUST_XLSX)
    gtot = gmatch = 0
    worst = 100
    sheets = SHEETS or [s for s in py if s in rs]
    for sh in sheets:
        p, r = py.get(sh, {}), rs.get(sh, {})
        tot = match = 0
        remaining = []
        for k in p:
            if k not in r:
                continue
            tot += 1
            if all(p[k][a] == r[k][a] for a in ("bold", "italic", "color")):
                match += 1
            else:
                remaining.append(k)
        pct = 100 * match // max(tot, 1)
        worst = min(worst, pct)
        gtot += tot
        gmatch += match
        print(f"{sh}: {match}/{tot} ({pct}%)  remaining={remaining[:10]}")
    print(f"OVERALL: {gmatch}/{gtot} ({100 * gmatch // max(gtot, 1)}%)")
    # Non-zero exit if any sheet drops below 100% — a real format-parity gate.
    return 0 if worst == 100 else 1


if __name__ == "__main__":
    sys.exit(main())
