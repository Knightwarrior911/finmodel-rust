"""Generate populated-IS parity oracles for the Rust IS-body port.

Unlike `build_excel_snapshots.py` (which runs with `is_structure=[]` → empty IS),
this passes a built `is_structure`, so the IS tab is fully populated. Inputs are
sourced from the committed empty-IS snapshots' `model_output` (the extraction
cache is gone), so the oracle's historicals match the Rust side exactly.

Output: tieout/excel_snapshots/<SAFE>_full_snapshot.json  (git-tracked oracle).
Run:    py tieout/build_full_is_oracle.py
"""
import json
import os
import sys
from pathlib import Path

os.environ["FINMODEL_DEV_MOCK"] = "1"
REPO = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(REPO))

SNAP_DIR = REPO / "tieout/excel_snapshots"
XLSX_DIR = REPO / "tests/snapshots"
XLSX_DIR.mkdir(parents=True, exist_ok=True)

COMPANIES = {
    "ASML_AS":   {"ticker": "ASML.AS",   "name": "ASML Holding NV",  "ccy": "EUR", "fye": "Sep"},
    "ATCO-B_ST": {"ticker": "ATCO-B.ST", "name": "Atlas Copco AB",   "ccy": "SEK", "fye": "Dec"},
    "NESN_SW":   {"ticker": "NESN.SW",   "name": "Nestle SA",        "ccy": "CHF", "fye": "Dec"},
    "NOVO-B_CO": {"ticker": "NOVO-B.CO", "name": "Novo Nordisk A/S", "ccy": "DKK", "fye": "Dec"},
    "SAND_ST":   {"ticker": "SAND.ST",   "name": "Sandvik AB",       "ccy": "SEK", "fye": "Dec"},
}


def characterize_xlsx(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            cells = []
            for cell in row:
                c = {"ref": cell.coordinate}
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        c["formula"] = cell.value
                    else:
                        c["value"] = cell.value
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                    c["fill"] = fill.fgColor.rgb
                if cell.hyperlink:
                    c["hyperlink"] = cell.hyperlink.target
                cells.append(c)
            if cells:
                rows.append({"row": row[0].row, "cells": cells})
        sheets[name] = rows
    wb.close()
    return sheets


def hist_slice(stmt: dict, nh: int) -> dict:
    return {k: (v[:nh] if isinstance(v, list) else v) for k, v in stmt.items()}


def build(safe: str, meta: dict, sector: str = "standard") -> None:
    from schemas.financial_data import ReconciledFinancialData, ModelConfig
    from src.engine import ModelEngine
    from src.writer import ExcelWriter
    from src.verifier import verify
    from src.assumptions import build_assumptions_block
    from src.is_builder import build_is_structure
    from src.utils import compute_historical_periods

    snap = json.loads((SNAP_DIR / f"{safe}_snapshot.json").read_text(encoding="utf-8"))
    mo_s = snap["model_output"]
    periods = snap["periods"]
    nh = sum(1 for p in periods if p.endswith("A"))
    hist_periods = [p for p in periods if p.endswith("A")]

    reconciled = ReconciledFinancialData(
        ticker=meta["ticker"], company_name=meta["name"], currency=meta["ccy"],
        fiscal_year_end=meta["fye"], periods=hist_periods,
        income_statement=hist_slice(mo_s["income_statement"], nh),
        balance_sheet=hist_slice(mo_s["balance_sheet"], nh),
        cash_flow_statement=hist_slice(mo_s["cash_flow_statement"], nh),
        notes={}, sources={}, flags=[],
    )
    cfg = ModelConfig(
        ticker=meta["ticker"], company_name=meta["name"], domicile="non-US",
        currency=meta["ccy"], fiscal_year_end=meta["fye"],
        periods_historical=nh, periods_projected=5, sector=sector,
    )
    derive_engine = ModelEngine(reconciled, cfg)
    hist = derive_engine._derive_assumptions()
    hp = compute_historical_periods(cfg.fiscal_year_end, cfg.periods_historical)
    ly = int(hp[-1][:4])
    pp = [f"{ly + i + 1}E" for i in range(cfg.periods_projected)]

    class _Stub:
        assumptions = hist
        periods = hp + pp

    assumptions = build_assumptions_block(_Stub(), cfg.ticker, sector=cfg.sector, reconciled=reconciled)
    engine = ModelEngine(reconciled, cfg, assumptions_block=assumptions)
    mo = engine.build()
    report = verify(mo, sector=cfg.sector)

    _is = mo.income_statement
    has_cogs = any(v and v != 0 for v in (_is.get("cogs") or []))
    has_rd = any(v and v != 0 for v in (_is.get("rd") or []))
    has_sga = any(v and v != 0 for v in (_is.get("sga") or []))
    is_structure = build_is_structure(cfg.sector, has_cogs=has_cogs, has_rd=has_rd, has_sga=has_sga)

    suffix = "_full" if sector == "standard" else f"_{sector}_full"
    xlsx = XLSX_DIR / f"{safe}{suffix}.xlsx"
    ExcelWriter(mo, report, meta["name"], str(xlsx), sources={}, currency=meta["ccy"],
                dcf=None, comps=None, assumptions=assumptions, ticker=meta["ticker"],
                fiscal_year_end=meta["fye"], sector=cfg.sector, is_structure=is_structure).write()

    sheets = characterize_xlsx(xlsx)
    out = {
        "ticker": meta["ticker"], "company": meta["name"], "currency": meta["ccy"],
        "periods": mo.periods, "sheets": sheets,
    }
    dst = SNAP_DIR / f"{safe}{suffix}_snapshot.json"
    dst.write_text(json.dumps(out, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"  {safe} [{sector}]: IS rows={len(sheets.get('IS', []))} -> {dst.name}")


def build_xbrl(safe: str, meta: dict) -> None:
    """Synthetic XBRL-detail oracle: 2 revenue segments + cogs_detail + rd/sga
    opex items forced onto the fixture (no US R-file fixtures exist). Model
    output is built manually (segment projection needs per-segment growth wiring
    the writer doesn't require — IS proj cells are formulas)."""
    from schemas.financial_data import ReconciledFinancialData, ModelConfig, ModelOutput
    from src.engine import ModelEngine
    from src.writer import ExcelWriter
    from src.verifier import verify
    from src.assumptions import build_assumptions_block
    from src.is_builder import build_is_structure
    from src.utils import compute_historical_periods

    snap = json.loads((SNAP_DIR / f"{safe}_snapshot.json").read_text(encoding="utf-8"))
    mo = snap["model_output"]; periods = snap["periods"]
    nh = sum(1 for p in periods if p.endswith("A")); n = len(periods)
    hh = lambda a: list(a[:nh])
    pad = lambda a: hh(a) + [None] * (n - nh)
    isd = {k: hh(v) for k, v in mo["income_statement"].items()}
    isd["rev_seg_a"] = [80000, 78000]; isd["rev_seg_b"] = [46503, 44878]
    isd["cogs_seg_a"] = [50000, 49000]; isd["cogs_seg_b"] = [24456, 24742]
    isd["opex_marketing"] = [3000, 3100]
    revenue_segments = [{"label": "Products", "key": "rev_seg_a"}, {"label": "Services", "key": "rev_seg_b"}]
    cogs_detail = [{"label": "Cost of products", "key": "cogs_seg_a"}, {"label": "Cost of services", "key": "cogs_seg_b"}]
    opex_items = [
        {"label": "Cost of products", "key": "cogs_seg_a", "category": "cogs", "group": ""},
        {"label": "Cost of services", "key": "cogs_seg_b", "category": "cogs", "group": ""},
        {"label": "Research & Development", "key": "rd", "category": "opex_rd", "group": ""},
        {"label": "Selling, General & Administrative", "key": "sga", "category": "opex", "group": ""},
        {"label": "Marketing", "key": "opex_marketing", "category": "opex", "group": ""},
    ]
    isd["cogs"] = list(isd["cogs_seg_a"]); isd["rd"] = list(isd["rd"]); isd["sga"] = list(isd["sga"])
    notes = {"revenue_segments": revenue_segments, "opex_items": opex_items, "cogs_detail": cogs_detail}
    rec = ReconciledFinancialData(
        ticker=meta["ticker"], company_name=meta["name"], currency=meta["ccy"],
        fiscal_year_end=meta["fye"], periods=[p for p in periods if p.endswith("A")],
        income_statement=isd, balance_sheet={k: hh(v) for k, v in mo["balance_sheet"].items()},
        cash_flow_statement={k: hh(v) for k, v in mo["cash_flow_statement"].items()},
        notes=notes, sources={}, flags=[])
    cfg = ModelConfig(ticker=meta["ticker"], company_name=meta["name"], domicile="non-US",
                      currency=meta["ccy"], fiscal_year_end=meta["fye"],
                      periods_historical=nh, periods_projected=5, sector="standard")
    cfg.revenue_segments = revenue_segments; cfg.cogs_detail = cogs_detail
    cfg.opex_items = opex_items; cfg.extra_opex_keys = [c["key"] for c in cogs_detail]
    cfg.nonrecurring_opex_keys = []
    de = ModelEngine(rec, cfg); hist = de._derive_assumptions()
    hp = compute_historical_periods(cfg.fiscal_year_end, nh); ly = int(hp[-1][:4])
    pp = [f"{ly + i + 1}E" for i in range(5)]

    class _Stub:
        assumptions = hist
        periods = hp + pp

    asmp = build_assumptions_block(_Stub(), cfg.ticker, sector="standard", reconciled=rec)
    mout = ModelOutput(periods=periods,
                       income_statement={k: pad(v) for k, v in isd.items()},
                       balance_sheet={k: pad(v) for k, v in mo["balance_sheet"].items()},
                       cash_flow_statement={k: pad(v) for k, v in mo["cash_flow_statement"].items()},
                       schedules={}, assumptions=hist, converged=True, plug_used=False)
    # seg_* keys trigger the "REVENUE BREAKDOWN BY SEGMENT" memo block (full
    # arrays so growth-scaled projections are exercised).
    mout.income_statement["revenue"] = list(mo["income_statement"]["revenue"])
    mout.income_statement["seg_NorthAmerica"] = [50000, 49000, 48000, 47000, 46000, 45000, 44000]
    mout.income_statement["seg_Europe"] = [76503, 73878, 71000, 68000, 65000, 62000, 59000]
    report = verify(mout, sector="standard")
    is_structure = build_is_structure("standard", has_cogs=True, has_rd=True, has_sga=True,
                                      revenue_segments=revenue_segments, opex_items=opex_items,
                                      cogs_detail=cogs_detail)
    xlsx = XLSX_DIR / f"{safe}_xbrl_full.xlsx"
    ExcelWriter(mout, report, meta["name"], str(xlsx), sources={}, currency=meta["ccy"],
                dcf=None, comps=None, assumptions=asmp, ticker=meta["ticker"],
                fiscal_year_end=meta["fye"], sector="standard", is_structure=is_structure).write()
    sheets = characterize_xlsx(xlsx)
    out = {
        "ticker": meta["ticker"], "company": meta["name"], "currency": meta["ccy"],
        "periods": mout.periods, "sheets": sheets,
        "model_output": {"income_statement": mout.income_statement,
                         "balance_sheet": mout.balance_sheet,
                         "cash_flow_statement": mout.cash_flow_statement},
    }
    dst = SNAP_DIR / f"{safe}_xbrl_full_snapshot.json"
    dst.write_text(json.dumps(out, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"  {safe} [xbrl]: IS rows={len(sheets.get('IS', []))} -> {dst.name}")


def main() -> None:
    runs = [(safe, meta, "standard") for safe, meta in COMPANIES.items()]
    # Sector coverage: force each non-standard sector on the SAND fixture so the
    # sector IS structures can be gated (no utility/bank/etc. fixtures exist).
    for sector in ("utility", "bank", "insurance", "reit"):
        runs.append(("SAND_ST", COMPANIES["SAND_ST"], sector))
    for safe, meta, sector in runs:
        try:
            build(safe, meta, sector)
        except Exception as e:
            import traceback
            print(f"  {safe} [{sector}]: FAILED {e}")
            traceback.print_exc()
    try:
        build_xbrl("SAND_ST", COMPANIES["SAND_ST"])
    except Exception as e:
        import traceback
        print(f"  SAND_ST [xbrl]: FAILED {e}")
        traceback.print_exc()


if __name__ == "__main__":
    main()
