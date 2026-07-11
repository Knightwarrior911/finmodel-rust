"""Generate the IFRS-16 bridge worksheet parity oracle for the Rust port.

Runs the reference Python `ResearchExcelWriter.write_ifrs_bridge` with two FIXED
inputs that exercise the branchy layout:
  * FULL  — IFRS→US GAAP, adjusted EBITDA present (differs from computed), EBITA
            section present, revenue>0 (margins).
  * SIMPLE— US GAAP→IFRS, no adjusted EBITDA (computed branch), no EBITA section,
            revenue==0 (no margins).
Reads each `.xlsx` back with openpyxl → cell-level snapshot (value/formula/fill).
Footer timestamp normalized (pinned) for determinism.

Run:  py tieout/build_ifrs_bridge_oracle.py
"""
import json
import sys
from pathlib import Path

REPO = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(REPO))

SNAP_DIR = REPO / "tieout/excel_snapshots"
XLSX_DIR = REPO / "tests/snapshots"
XLSX_DIR.mkdir(parents=True, exist_ok=True)
SNAP_DIR.mkdir(parents=True, exist_ok=True)

PINNED_GENERATED = "Generated: 2026-01-01 00:00 | Source: SEC EDGAR / yfinance / Company filings"


def characterize_xlsx(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            cells = []
            for cell in row:
                c = {"ref": cell.coordinate}
                if cell.value is not None:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        c["formula"] = val
                    else:
                        if isinstance(val, str) and val.startswith("Generated: "):
                            val = PINNED_GENERATED
                        c["value"] = val
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                    c["fill"] = fill.fgColor.rgb
                cells.append(c)
            if cells:
                rows.append({"row": row[0].row, "cells": cells})
        sheets[name] = rows
    wb.close()
    return sheets


def cases():
    from kb.ifrs import IFRSAdjustmentInput, auto_convert

    full_in = IFRSAdjustmentInput(
        rou_depreciation=80.0,
        lease_interest=20.0,
        short_term_rent=30.0,
        reported_ebit=1000.0,
        reported_ebitda=1400.0,   # > EBIT and differs from computed 1250 → adjusted branch
        reported_ebita=1100.0,    # != EBIT → EBITA section
        standard_depreciation=200.0,
        standard_amortization=50.0,
        accounting_standard="IFRS",
    )
    full_out = auto_convert(full_in, 5000.0)

    simple_in = IFRSAdjustmentInput(
        rou_depreciation=40.0,
        lease_interest=10.0,
        short_term_rent=0.0,
        reported_ebit=500.0,
        reported_ebitda=0.0,      # not > EBIT → computed branch (no adjusted)
        reported_ebita=500.0,     # == EBIT → no EBITA section
        standard_depreciation=100.0,
        standard_amortization=0.0,
        accounting_standard="US GAAP",
    )
    simple_out = auto_convert(simple_in, 0.0)

    return [
        ("IFRS_BRIDGE", "IFRS_BRIDGE.xlsx", full_in, full_out, "FullCo", "FY2024", 5000.0),
        ("IFRS_BRIDGE_SIMPLE", "IFRS_BRIDGE_SIMPLE.xlsx", simple_in, simple_out, "SimpleCo", "FY2024", 0.0),
    ]


def main() -> None:
    from src.research.output_writer import ResearchExcelWriter

    for tag, fname, inp, out, company, period, revenue in cases():
        writer = ResearchExcelWriter(output_dir=str(XLSX_DIR))
        writer.write_ifrs_bridge(inp, out, company, period, revenue=revenue, filename=fname)
        sheets = characterize_xlsx(XLSX_DIR / fname)
        dst = SNAP_DIR / f"{tag}_snapshot.json"
        dst.write_text(json.dumps({"sheets": sheets}, indent=2), encoding="utf-8")
        n_cells = sum(len(c["cells"]) for rs in sheets.values() for c in rs)
        print(f"  {tag}: rows={sum(len(rs) for rs in sheets.values())} "
              f"cells={n_cells} -> {dst.name}")


if __name__ == "__main__":
    main()
