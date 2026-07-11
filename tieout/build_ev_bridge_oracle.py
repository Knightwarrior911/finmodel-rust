"""Generate the EV-bridge worksheet parity oracle for the Rust port.

Runs the reference Python `ResearchExcelWriter.write_ev_bridge` with a FIXED
`EVBridgeInput` (all add/sub items present + valuation multiples), reads the
`.xlsx` back with openpyxl, and dumps a cell-level snapshot
(`value`/`formula`/`fill`) to
`tieout/excel_snapshots/EV_BRIDGE_snapshot.json`. The footer timestamp is
normalized (pinned) for determinism; the Rust gate uses the same string.

Run:  py tieout/build_ev_bridge_oracle.py
"""
import json
import sys
from pathlib import Path

REPO = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(REPO))

SNAP_DIR = REPO / "tieout/excel_snapshots"
XLSX_DIR = REPO / "tests/snapshots"
XLSX_DIR.mkdir(parents=True, exist_ok=True)
SNAP_DIR.mkdir(parents=True, exist_ok=True)

PINNED_GENERATED = "Generated: 2026-01-01 00:00 | Source: SEC EDGAR / yfinance / Company filings"


def characterize_xlsx(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            cells = []
            for cell in row:
                c = {"ref": cell.coordinate}
                if cell.value is not None:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        c["formula"] = val
                    else:
                        if isinstance(val, str) and val.startswith("Generated: "):
                            val = PINNED_GENERATED
                        c["value"] = val
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                    c["fill"] = fill.fgColor.rgb
                cells.append(c)
            if cells:
                rows.append({"row": row[0].row, "cells": cells})
        sheets[name] = rows
    wb.close()
    return sheets


def fixed_input():
    from kb.ev_bridge import EVBridgeInput
    return EVBridgeInput(
        company="DemoCo",
        currency="USD",
        share_price=150.0,
        shares_outstanding=1_000_000_000.0,
        total_debt=50_000_000_000.0,
        finance_leases=5_000_000_000.0,
        operating_leases=8_000_000_000.0,
        underfunded_pension=2_000_000_000.0,
        minority_interest=1_000_000_000.0,
        preferred_stock=500_000_000.0,
        cash=20_000_000_000.0,
        short_term_investments=10_000_000_000.0,
        equity_investments=3_000_000_000.0,
        nol_dta=1_500_000_000.0,
        ltm_revenue=100_000_000_000.0,
        ltm_ebitda=30_000_000_000.0,
    )


def main() -> None:
    from src.research.output_writer import ResearchExcelWriter
    from kb.ev_bridge import EVBridgeInput

    # Full case — every add/sub item + both multiples.
    cases = [("EV_BRIDGE", "EV_BRIDGE.xlsx", fixed_input())]
    # Sparse case — several add/sub items absent (exercises the dynamic row
    # shifts in the EV formula + multiples row-refs). Revenue present so the
    # EV/EBITDA formula uses the r_rev+1 (correct) branch.
    sparse = EVBridgeInput(
        company="SparseCo",
        currency="USD",
        share_price=42.0,
        shares_outstanding=500_000_000.0,
        total_debt=12_000_000_000.0,
        minority_interest=800_000_000.0,
        cash=6_000_000_000.0,
        nol_dta=400_000_000.0,
        ltm_revenue=25_000_000_000.0,
        ltm_ebitda=4_000_000_000.0,
    )
    cases.append(("EV_BRIDGE_SPARSE", "EV_BRIDGE_SPARSE.xlsx", sparse))

    for tag, fname, inp in cases:
        writer = ResearchExcelWriter(output_dir=str(XLSX_DIR))
        writer.write_ev_bridge(inp, filename=fname)
        sheets = characterize_xlsx(XLSX_DIR / fname)
        dst = SNAP_DIR / f"{tag}_snapshot.json"
        dst.write_text(json.dumps({"sheets": sheets}, indent=2), encoding="utf-8")
        n_cells = sum(len(c["cells"]) for rs in sheets.values() for c in rs)
        print(f"  {tag}: rows={sum(len(rs) for rs in sheets.values())} "
              f"cells={n_cells} -> {dst.name}")


if __name__ == "__main__":
    main()
