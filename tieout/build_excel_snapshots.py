"""Task 0.5.1 — Excel characterization snapshots from cached baseline extractions.

Builds .xlsx models + full cell-level .json snapshots for the 5 baseline
European companies. The Rust port's parity gate (R.5/R.6) must reproduce
both cell values AND formulas AND tier colors AND hyperlinks.

Outputs:
  tests/snapshots/          ← git-ignored; generated .xlsx files
  tieout/excel_snapshots/   ← git-tracked; per-cell characterization .json
"""
import json
import os
import sys
from pathlib import Path

os.environ["FINMODEL_DEV_MOCK"] = "1"
REPO = Path(__file__).parent.resolve()
sys.path.insert(0, str(REPO))

COMPANIES = {
    "ASML_AS":   {"ticker": "ASML.AS",  "name": "ASML Holding NV",     "ccy": "EUR", "fye": "Sep"},
    "ATCO-B_ST": {"ticker": "ATCO-B.ST", "name": "Atlas Copco AB",      "ccy": "SEK", "fye": "Dec"},
    "NESN_SW":   {"ticker": "NESN.SW",   "name": "Nestle SA",           "ccy": "CHF", "fye": "Dec"},
    "NOVO-B_CO": {"ticker": "NOVO-B.CO",  "name": "Novo Nordisk A/S",   "ccy": "DKK", "fye": "Dec"},
    "SAND_ST":   {"ticker": "SAND.ST",   "name": "Sandvik AB",          "ccy": "SEK", "fye": "Dec"},
}

FP = "4065a2c76ef95ca6"
CACHE_DIR = REPO / "tieout/results/_modelcache"
XLSX_DIR = REPO / "tests/snapshots"
SNAPSHOT_DIR = REPO / "tieout/excel_snapshots"
XLSX_DIR.mkdir(parents=True, exist_ok=True)
SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)


def safe(ticker: str) -> str:
    return ticker.replace(".", "_").replace("/", "_")


def load_cache(ticker_safe: str) -> dict:
    return json.loads((CACHE_DIR / f"{FP}_{ticker_safe}.json").read_text(encoding="utf-8"))


def characterize_xlsx(path: Path) -> dict:
    """Read xlsx with openpyxl; capture values, formulas, fill colors, hyperlinks per sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            cells = []
            for cell in row:
                c = {"ref": cell.coordinate}
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        c["formula"] = cell.value
                    else:
                        c["value"] = cell.value
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                    c["fill"] = fill.fgColor.rgb
                if cell.hyperlink:
                    c["hyperlink"] = cell.hyperlink.target
                cells.append(c)
            if cells:
                rows.append({"row": row[0].row, "cells": cells})
        sheets[name] = rows
    wb.close()
    return sheets


def build_snapshot(ticker_safe: str, meta: dict):
    from schemas.financial_data import ReconciledFinancialData, ModelConfig
    from src.engine import ModelEngine
    from src.writer import ExcelWriter
    from src.verifier import verify

    cache = load_cache(ticker_safe)
    years = cache.get("years_found", [])
    periods = [f"{y}A" for y in years]

    reconciled = ReconciledFinancialData(
        ticker=meta["ticker"], company_name=meta["name"],
        currency=meta["ccy"], fiscal_year_end=meta["fye"],
        periods=periods,
        income_statement=cache.get("income_statement", {}),
        balance_sheet=cache.get("balance_sheet", {}),
        cash_flow_statement=cache.get("cash_flow_statement", {}),
        notes={}, sources={}, flags=[],
    )

    cfg = ModelConfig(
        ticker=meta["ticker"], company_name=meta["name"],
        domicile="non-US", currency=meta["ccy"],
        fiscal_year_end=meta["fye"],
        periods_historical=len(periods), periods_projected=5,
        sector="standard",
    )

    derive_engine = ModelEngine(reconciled, cfg)
    hist = derive_engine._derive_assumptions()

    from src.utils import compute_historical_periods
    hp = compute_historical_periods(cfg.fiscal_year_end, cfg.periods_historical)
    ly = int(hp[-1][:4])
    pp = [f"{ly + i + 1}E" for i in range(cfg.periods_projected)]

    class _Stub:
        assumptions = hist
        periods = hp + pp

    from src.assumptions import build_assumptions_block
    assumptions = build_assumptions_block(_Stub(), cfg.ticker, sector=cfg.sector,
                                          reconciled=reconciled)

    engine = ModelEngine(reconciled, cfg, assumptions_block=assumptions)
    mo = engine.build()
    report = verify(mo, sector=cfg.sector)

    xlsx_path = XLSX_DIR / f"{safe(meta['ticker'])}_snapshot.xlsx"
    ExcelWriter(mo, report, meta["name"], str(xlsx_path),
                sources=reconciled.sources, currency=meta["ccy"],
                dcf=None, comps=None, assumptions=assumptions,
                ticker=meta["ticker"], sector=cfg.sector).write()
    print(f"  xlsx -> {xlsx_path}")

    # Full cell-level characterization
    sheets = characterize_xlsx(xlsx_path)

    snap = {
        "ticker": meta["ticker"],
        "company": meta["name"],
        "currency": meta["ccy"],
        "periods": mo.periods,
        "converged": mo.converged,
        "extraction_years": years,
        "source_fingerprint": FP,
        "model_output": {
            "income_statement": dict(mo.income_statement),
            "balance_sheet": dict(mo.balance_sheet),
            "cash_flow_statement": dict(mo.cash_flow_statement),
        },
        "sheets": sheets,
        "assumptions": {
            "revenue_growth_pct": assumptions.base.revenue_growth_pct,
            "gross_margin_pct": assumptions.base.gross_margin_pct,
            "active_case": assumptions.active_case,
        },
        "verification": {
            "passed": report.passed,
            "critical_failures": report.critical_failures,
            "warnings": report.warnings,
        },
    }
    sp = SNAPSHOT_DIR / f"{safe(meta['ticker'])}_snapshot.json"
    sp.write_text(json.dumps(snap, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"  json -> {sp} ({sp.stat().st_size:,} bytes)")


def main():
    for k, meta in COMPANIES.items():
        try:
            print(f"\n=== {meta['ticker']} ===")
            build_snapshot(k, meta)
        except Exception as e:
            print(f"  FAIL: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    main()
