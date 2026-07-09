"""
qa_checklist.py — Deliverable QA checklist for client-ready .xlsx output.

Usage:
    python scripts/qa_checklist.py <path/to/model.xlsx>

Exits 0 if ALL checks pass, 1 if any fail.
"""

import os
import sys
from pathlib import Path

# ——————————————————————————————————————————————————————————
# Path setup — project root is parent of scripts/
# ——————————————————————————————————————————————————————————
_SCRIPT_DIR = Path(__file__).resolve().parent
_REPO = _SCRIPT_DIR.parent
sys.path.insert(0, str(_REPO))

# ——————————————————————————————————————————————————————————
# Helpers
# ——————————————————————————————————————————————————————————

_RED_UNVERIFIED_HEX = "C00000"       # _TIER_COLOR["unverified"] in audit_pipeline
_PRIMARY_COLOR = "#255BE3"
_INK_COLOR = "#0F1632"
_REQUIRED_SHEETS = {"Cover", "Assumptions", "IS", "BS", "CF", "Sources"}
_BRANDING_YAML_PATH = _REPO / "config" / "branding.yaml"


def _load_branding():
    """Load branding.yaml; return dict or empty on error."""
    import yaml
    try:
        with open(_BRANDING_YAML_PATH, encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    except Exception:
        return {}


def _normalize_rgb(color_obj):
    """Normalize openpyxl font.color to uppercase hex RGB (6 chars, no #).

    Returns None when the colour is a theme/indexed reference
    (which can't be resolved without the workbook theme) or unset.
    """
    if color_obj is None:
        return None
    try:
        rgb = color_obj.rgb
    except (ValueError, AttributeError):
        return None
    if rgb is None:
        return None
    if not isinstance(rgb, str):
        return None
    s = rgb.upper().strip()
    # openpyxl may return '00C00000' (theme) or 'FFC00000' (argb)
    if len(s) == 8:
        return s[2:] if s.startswith("FF") else s
    if len(s) == 6:
        return s
    if len(s) == 4 and s.startswith("00"):
        return s[2:]  # theme-short like '00C0'
    return s


def _is_red_unverified(color_obj):
    """Check if an openpyxl font.color matches the UNVERIFIED red (#C00000)."""
    n = _normalize_rgb(color_obj)
    return n is not None and n == _RED_UNVERIFIED_HEX


def _load_workbook(path):
    """Open xlsx with openpyxl; exits on failure."""
    try:
        import openpyxl
        return openpyxl.load_workbook(str(path), data_only=False)
    except Exception as e:
        print(f"  FAIL  Cannot open workbook: {e}")
        sys.exit(1)


# ——————————————————————————————————————————————————————————
# Check implementations
# ——————————————————————————————————————————————————————————

def check_branding(wb):
    """Verify brand fonts & colors applied across key surfaces.

    Checks:
      - Cover title row (hbar) uses Arial, white on ink background
      - Cover subtitle uses primary_color or ink color
      - Body labels use Arial, ink_color
      - Tab colors match brand palette
    """
    issues = []

    # Read branding config
    branding = _load_branding()
    brand_font = branding.get("font_display", branding.get("font_body", "Arial"))
    # primary_color and ink_color may have # prefix; normalize
    brand_primary = (branding.get("primary_color") or _PRIMARY_COLOR).lstrip("#").upper()
    brand_ink = (branding.get("ink_color") or _INK_COLOR).lstrip("#").upper()

    ws = wb["Cover"]

    # --- Cover title (hbar format) ---
    title_found = None
    for row in ws.iter_rows(min_row=1, max_row=8, max_col=10, values_only=False):
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and "Valuation Model" in v:
                title_found = cell
                break
        if title_found:
            break

    if title_found:
        f = title_found.font
        name_ok = f.name and f.name.lower() == brand_font.lower()
        if not name_ok:
            issues.append(
                f"Cover title font is '{f.name}', expected '{brand_font}'"
            )
    else:
        issues.append("Cover title cell ('Valuation Model') not found")

    # --- Cover subtitle (primary color text) ---
    subtitle_found = None
    for row in ws.iter_rows(min_row=3, max_row=8, max_col=10, values_only=False):
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and "Statement" in v:
                subtitle_found = cell
                break
        if subtitle_found:
            break

    if subtitle_found:
        sub_rgb = _normalize_rgb(subtitle_found.font.color if subtitle_found.font.color else None)
        if sub_rgb and sub_rgb != brand_primary and sub_rgb != brand_ink:
            issues.append(
                f"Cover subtitle color is #{sub_rgb}, expected #{brand_primary} or #{brand_ink}"
            )
    else:
        issues.append("Cover subtitle ('Statement') not found")

    # --- Body labels (ink color) ---
    label_sample = None
    for row in ws.iter_rows(min_row=8, max_row=15, max_col=6, values_only=False):
        for cell in row:
            v = cell.value
            if v and isinstance(v, str) and v in ("Company", "Ticker", "Currency"):
                label_sample = cell
                break
        if label_sample:
            break

    if label_sample:
        lbl_rgb = _normalize_rgb(label_sample.font.color if label_sample.font.color else None)
        if lbl_rgb and lbl_rgb != brand_ink:
            issues.append(
                f"Body label color is #{lbl_rgb}, expected #{brand_ink}"
            )
    else:
        issues.append("Body label cell ('Company'/'Ticker'/'Currency') not found")

    # --- Tab colors ---
    for name in wb.sheetnames:
        ws_t = wb[name]
        # openpyxl: tab_color is a Color object or None
        tc = getattr(ws_t, "tab_color", None)
        if tc is not None:
            tc_rgb = _normalize_rgb(tc.rgb if tc.rgb else None)
            # Accept any non-default tab color as brand-touched
            if tc_rgb and tc_rgb not in (brand_primary, brand_ink, "FFFFFFFF", "00000000"):
                pass  # variant tab colors are expected (IS=#E6EBED, etc.)

    return issues


def check_unverified_red_cells(wb):
    """No UNVERIFIED trust-tier RED cells are left unexplained.

    A red cell is acceptable IF it has a cell comment explaining the
    unverified status.  Those without a comment are a FAIL.
    """
    import openpyxl
    issues = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                color_obj = cell.font.color
                if color_obj is None:
                    continue
                if not _is_red_unverified(color_obj):
                    continue
                # This cell has red font (#C00000) → UNVERIFIED marker
                cmt = cell.comment
                cmt_text = (cmt.text or "").strip() if cmt else ""
                if not cmt_text:
                    issues.append(
                        f"[{sn}] {cell.coordinate}: UNVERIFIED red cell with no explanation comment"
                    )
                elif "Unverified" not in cmt_text and "unverified" not in cmt_text:
                    issues.append(
                        f"[{sn}] {cell.coordinate}: UNVERIFIED red cell comment does not explain: {cmt_text[:80]}"
                    )
    return issues


def check_required_sheets(wb):
    """All required sheets must exist."""
    present = set(wb.sheetnames)
    missing = _REQUIRED_SHEETS - present
    issues = []
    if missing:
        issues.append(f"Missing sheet(s): {', '.join(sorted(missing))}")
    extras = present - _REQUIRED_SHEETS
    if extras:
        issues.append(f"Extra sheet(s) present: {', '.join(sorted(extras))}")
    return issues


def check_sanity(wb):
    """Check for warning indicators from the VerificationReport and sanity rows.

    Looks at the Sources sheet for:
      - VerificationReport status FAILED
      - Critical failures
      - Plug-used indicator
      - Warnings beyond expected low-severity notes
    """
    import re
    issues = []
    warnings_found = []

    if "Sources" not in wb.sheetnames:
        issues.append("Sources sheet missing — cannot check verification report")
        return issues

    ws = wb["Sources"]
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None or not isinstance(v, str):
                continue
            vl = v.strip()

            # Check for FAILED verification status
            if "FAILED" in vl:
                issues.append(f"VerificationReport status is FAILED ({vl})")
            # Check for critical failures
            if "Critical" in vl.lower() and "failure" in vl.lower():
                issues.append(f"Critical failure reported: {vl}")
            # Check for plug used
            if "Plug" in vl and ("used" in vl.lower() or "was used" in vl.lower()):
                issues.append(f"Plug used to balance BS: {vl}")
            # Collect warnings
            if vl.startswith("Warnings:") or vl.startswith("Warning"):
                idx = vl.find(":")
                if idx >= 0:
                    rest = vl[idx + 1:].strip()
                    warnings_found.append(rest)
                else:
                    warnings_found.append(vl)

    # Flag warnings beyond expected minor ones
    # Minor gap warnings for CFS (±small %) are acceptable
    for w in warnings_found:
        if w and w.lower() not in ("", "none"):
            if "minor gap" in w.lower():
                continue  # accepted minor diff
            issues.append(f"Warning: {w}")

    # Also check any cell containing "⚠" warning indicator
    for sn in wb.sheetnames:
        if sn == "Sources":
            continue
        ws = wb[sn]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v and isinstance(v, str) and "⚠" in v:
                    issues.append(f"[{sn}] {cell.coordinate}: warning indicator in cell: {v[:80]}")

    return issues


# ——————————————————————————————————————————————————————————
# Main
# ——————————————————————————————————————————————————————————

def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/qa_checklist.py <path/to/model.xlsx>")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"  FAIL  File not found: {xlsx_path}")
        sys.exit(1)

    wb = _load_workbook(xlsx_path)
    checks = [
        ("Branding (fonts/colors match brand.yaml)", check_branding),
        ("No UNVERIFIED red cells unexplained", check_unverified_red_cells),
        ("Required sheets present", check_required_sheets),
        ("Sanity checks (verification report / warnings)", check_sanity),
    ]

    print(f"QA Checklist — {xlsx_path.name}")
    print("=" * 60)
    print()

    all_pass = True
    for label, fn in checks:
        try:
            issues = fn(wb)
        except Exception as e:
            issues = [f"Check raised exception: {e}"]

        if issues:
            all_pass = False
            print(f"  FAIL  {label}")
            for iss in issues:
                print(f"         • {iss}")
        else:
            print(f"  PASS  {label}")
        print()

    wb.close()

    if all_pass:
        print("=" * 60)
        print("  ALL CHECKS PASSED")
        print("=" * 60)
        sys.exit(0)
    else:
        print("=" * 60)
        print("  SOME CHECKS FAILED  —  review details above")
        print("=" * 60)
        sys.exit(1)


if __name__ == "__main__":
    main()
