"""Tests for src.audit_pipeline — provenance attach + file#page xlsx links."""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest

from src.audit_pipeline import (
    attach_provenance_to_cache,
    annotate_workbook_with_links,
    run_audit,
)


def _atco_pdf() -> Path:
    p = Path("extraction_cache/ATCO_B_2023_raw.pdf")
    if not p.exists():
        pytest.skip(f"ATCO PDF not present at {p}")
    return p


def _atco_cache() -> Path:
    p = Path("extraction_cache/ATCO-B_ST.json")
    if not p.exists():
        pytest.skip(f"ATCO cache not present at {p}")
    return p


def test_attach_provenance_to_cache(tmp_path):
    """Copy cache + run attach; assert __provenance__ key populated."""
    cache_src = _atco_cache()
    pdf = _atco_pdf()
    cache = tmp_path / "ATCO-B_ST.json"
    shutil.copy(cache_src, cache)

    n_located = attach_provenance_to_cache(cache, pdf)
    assert n_located > 0, "should locate at least one value"

    out = json.loads(cache.read_text(encoding="utf-8"))
    assert "__provenance__" in out
    assert "income_statement.revenue" in out["__provenance__"]


def test_annotate_workbook_with_links(tmp_path):
    """Build a fake xlsx whose cell value matches a cached ATCO number,
    run annotate, assert a file#page hyperlink was added."""
    import openpyxl
    cache_src = _atco_cache()
    pdf = _atco_pdf()
    cache = tmp_path / "ATCO-B_ST.json"
    shutil.copy(cache_src, cache)
    attach_provenance_to_cache(cache, pdf)

    sample_cache = json.loads(cache.read_text(encoding="utf-8"))
    rev = sample_cache["income_statement"]["revenue"][0]   # 2023 = 172664
    xlsx = tmp_path / "fake_model.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Revenue 2023"
    ws["B1"] = rev
    wb.save(str(xlsx))

    annotated = annotate_workbook_with_links(xlsx, cache_path=cache)
    assert annotated["total"] >= 1

    wb2 = openpyxl.load_workbook(str(xlsx))
    cell = wb2.active["B1"]
    assert cell.hyperlink is not None
    target = cell.hyperlink.target
    assert target.startswith("file:///")
    assert "#page=" in target


def test_run_audit_end_to_end(tmp_path):
    """Run full audit pipeline against ATCO-B.ST, annotating a fake xlsx."""
    import openpyxl
    cache_dir = tmp_path / "extraction_cache"
    cache_dir.mkdir()
    shutil.copy(_atco_cache(), cache_dir / "ATCO-B_ST.json")
    shutil.copy(_atco_pdf(), cache_dir / "ATCO_B_2023_raw.pdf")

    # A workbook carrying the 2023 revenue value
    cache = json.loads((cache_dir / "ATCO-B_ST.json").read_text(encoding="utf-8"))
    rev = cache["income_statement"]["revenue"][0]
    xlsx = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "Revenue"; ws["B1"] = rev
    wb.save(str(xlsx))

    res = run_audit("ATCO-B.ST", cache_dir=cache_dir, xlsx_path=xlsx)
    assert res["ok"] is True, res
    assert res["values_located"] > 0
    assert res["annotated"]["total"] > 0
    # No snapshots artifacts produced anywhere.
    assert not (tmp_path / "snapshots").exists()


def test_run_audit_no_pdf_returns_error(tmp_path):
    cache_dir = tmp_path / "extraction_cache"
    cache_dir.mkdir()
    shutil.copy(_atco_cache(), cache_dir / "ATCO-B_ST.json")
    # NB: no PDF copied

    res = run_audit("ATCO-B.ST", cache_dir=cache_dir)
    assert res["ok"] is False
    assert "PDF not found" in res["error"]
