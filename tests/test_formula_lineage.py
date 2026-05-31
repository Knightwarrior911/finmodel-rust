import json
import openpyxl
from src.source_ledger import SourceLedger
from src.audit_pipeline import annotate_workbook_with_links


def _wb(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IS"
    ws["A1"] = "Revenue"; ws["B1"] = 1000
    ws["A2"] = "COGS"; ws["B2"] = 600
    ws["A3"] = "Gross Profit"; ws["B3"] = "=B1-B2"
    p = tmp_path / "m.xlsx"; wb.save(p); return p


def _cache_with_ledger(tmp_path):
    led = SourceLedger()
    led.record_derived("x", "y", None, value=1.0, formula="f", inputs=[])
    p = tmp_path / "c.json"
    p.write_text(json.dumps({"__ledger__": led.to_json()}), encoding="utf-8")
    return p


def test_formula_cell_gets_lineage_comment(tmp_path):
    xlsx = _wb(tmp_path)
    cp = _cache_with_ledger(tmp_path)
    res = annotate_workbook_with_links(str(xlsx), cache_path=str(cp))
    assert res["derived_formula"] >= 1
    assert "covered_pct" in res
    wb = openpyxl.load_workbook(str(xlsx))
    c = wb["IS"]["B3"]
    assert c.comment is not None
    assert "Computed" in c.comment.text
    assert "Revenue" in c.comment.text and "COGS" in c.comment.text


def test_no_ledger_leaves_formula_uncommented(tmp_path):
    xlsx = _wb(tmp_path)
    cp = tmp_path / "empty.json"; cp.write_text("{}", encoding="utf-8")
    annotate_workbook_with_links(str(xlsx), cache_path=str(cp))
    wb = openpyxl.load_workbook(str(xlsx))
    assert wb["IS"]["B3"].comment is None
