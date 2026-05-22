"""Tests for src.citations — market-data citations + their xlsx linking."""
from __future__ import annotations

import json

from src.citations import (
    MarketCitation,
    yahoo_url,
    collect_market_citations,
    persist_citations,
    load_citations,
)


def test_yahoo_url_encodes_caret():
    assert yahoo_url("^TNX") == "https://finance.yahoo.com/quote/%5ETNX"
    assert yahoo_url("MSFT") == "https://finance.yahoo.com/quote/MSFT"


def test_collect_from_public_comps():
    from schemas.financial_data import PublicCompPeer, PublicCompsOutput
    peer = PublicCompPeer(ticker="MSFT", name="Microsoft",
                          share_price=445.0, market_cap=1068000.0,
                          ltm_revenue=245000.0)
    out = PublicCompsOutput(target_ticker="AAPL", target_company_name="Apple",
                            as_of_date="2026-05-22", peers=[peer])
    cites = collect_market_citations(public_comps=out, target_ticker="AAPL")
    by_val = {c.value: c for c in cites}
    assert 1068000.0 in by_val
    assert by_val[1068000.0].url == "https://finance.yahoo.com/quote/MSFT"
    assert "MSFT" in by_val[1068000.0].label


def test_collect_wacc_risk_free():
    from schemas.financial_data import WACCOutput
    w = WACCOutput(
        peers=[], median_unlevered_beta=0.9, target_levered_beta=1.1,
        target_de_ratio=0.3, risk_free_rate=0.045, equity_risk_premium=0.055,
        cost_of_equity=0.10, cost_of_debt_pretax=0.035, tax_rate=0.21,
        after_tax_cost_of_debt=0.028, target_market_cap=500000.0,
        target_debt=30000.0, target_total_capital=530000.0,
        equity_weight=0.94, debt_weight=0.06, wacc=0.10,
    )
    cites = collect_market_citations(wacc=w, target_ticker="ATCO-B.ST")
    vals = {round(c.value, 4) for c in cites}
    assert 0.045 in vals          # risk-free rate cited
    assert 500000.0 in {c.value for c in cites}   # target market cap cited


def test_persist_and_load_roundtrip(tmp_path):
    cache = tmp_path / "X.json"
    cache.write_text(json.dumps({"years_found": ["2024"]}), encoding="utf-8")
    cites = [MarketCitation(1068000.0, "MSFT market cap",
                            "https://finance.yahoo.com/quote/MSFT", "yfinance",
                            "2026-05-22")]
    n = persist_citations(cache, cites)
    assert n == 1
    data = json.loads(cache.read_text(encoding="utf-8"))
    assert "__citations__" in data
    assert data["years_found"] == ["2024"]            # other keys preserved
    back = load_citations(data["__citations__"])
    assert back[0].value == 1068000.0
    assert back[0].url.endswith("/MSFT")


def test_annotate_links_market_value(tmp_path):
    """A cell whose value matches a market citation gets the provider URL."""
    import openpyxl
    from src.audit_pipeline import annotate_workbook_with_links

    cache = tmp_path / "X.json"
    cache.write_text(json.dumps({
        "years_found": ["2024"],
        "__citations__": [MarketCitation(
            1068000.0, "MSFT market cap",
            "https://finance.yahoo.com/quote/MSFT", "yfinance", "2026-05-22"
        ).to_json()],
    }), encoding="utf-8")

    xlsx = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "MSFT"; ws["B1"] = 1068000.0
    wb.save(str(xlsx))

    res = annotate_workbook_with_links(xlsx, cache_path=cache)
    assert res["linked_market"] == 1

    wb2 = openpyxl.load_workbook(str(xlsx))
    assert wb2.active["B1"].hyperlink.target == "https://finance.yahoo.com/quote/MSFT"
