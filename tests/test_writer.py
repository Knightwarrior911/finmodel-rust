# financial_model/tests/test_writer.py
import os
import tempfile
import pytest
import xlsxwriter
from pathlib import Path
from schemas.financial_data import ModelOutput, VerificationReport
from src.writer import ExcelWriter

PERIODS = ["2021A", "2022A", "2023A", "2024E", "2025E", "2026E"]
N = len(PERIODS)

SAMPLE_OUTPUT = ModelOutput(
    periods=PERIODS,
    income_statement={
        "revenue": [365817, 394328, 383285, 400000, 416000, 432640],
        "cogs": [212981, 223546, 214137, 224000, 233000, 242000],
        "gross_profit": [152836, 170782, 169148, 176000, 183000, 190640],
        "sga": [21973, 25094, 24932, 26000, 27000, 28000],
        "rd": [21914, 26251, 29915, 30000, 31000, 32000],
        "da": [11284, 11104, 11519, 12000, 12480, 12979],
        "ebit": [108949, 119437, 114301, 118000, 123000, 128000],
        "interest_expense": [2645, 2830, 3933, 3500, 3400, 3300],
        "interest_income": [2843, 2825, 3750, 2000, 2100, 2200],
        "income_tax": [14527, 19300, 16520, 17000, 18000, 19000],
        "net_income": [94680, 99803, 96995, 99500, 103700, 107900],
        "eps_diluted": [5.61, 6.15, 6.13, 6.50, 6.90, 7.35],
        "eps_basic": [5.67, 6.23, 6.16, 6.55, 6.95, 7.40],
        "shares_diluted": [16865, 16215, 15813, 15300, 15000, 14700],
        "shares_basic": [16701, 16030, 15744, 15200, 14900, 14600],
    },
    balance_sheet={
        "cash": [62639, 48304, 61555, 65000, 68000, 71000],
        "accounts_receivable": [26278, 28184, 29508, 30000, 31000, 32000],
        "inventory": [6580, 4946, 6331, 6500, 6700, 6900],
        "total_current_assets": [134836, 135405, 143566, 148000, 152000, 156000],
        "ppe_net": [39440, 42117, 43715, 44000, 44500, 45000],
        "goodwill": [0, 0, 0, 0, 0, 0],
        "intangibles_net": [0, 0, 0, 0, 0, 0],
        "total_assets": [351002, 352755, 352583, 358000, 362000, 366000],
        "accounts_payable": [54763, 64115, 62611, 63000, 64000, 65000],
        "total_current_liabilities": [125481, 153982, 145308, 148000, 151000, 154000],
        "long_term_debt": [109106, 98959, 95281, 90000, 85000, 80000],
        "total_liabilities": [287912, 302083, 290437, 292000, 294000, 295000],
        "retained_earnings": [5562, -3068, -214, 5000, 10000, 15000],
        "total_equity": [63090, 50672, 62146, 66000, 68000, 71000],
    },
    cash_flow_statement={
        "cfo": [104038, 122151, 110543, 112000, 115000, 118000],
        "capex": [11085, 10708, 10959, 11500, 12000, 12500],
        "cfi": [-14545, -22354, -3, -12000, -12500, -13000],
        "cff": [-93353, -110749, -108488, -100000, -102000, -104000],
        "net_change_cash": [-3860, -10952, 13248, 0, 0, 0],
    },
    schedules={
        "ppe_rollforward": [
            {"period": p, "opening": 39000, "capex": 11000, "da": 11000, "closing": 39000}
            for p in PERIODS
        ]
    },
    assumptions={
        "revenue_growth_pct": 0.041,
        "gross_margin_pct": 0.44,
        "sga_pct_rev": 0.065,
        "rd_pct_rev": 0.074,
        "da_pct_rev": 0.030,
        "capex_pct_rev": 0.028,
        "tax_rate_pct": 0.155,
        "interest_rate_pct": 0.035,
        "dso_days": 28.0,
        "dpo_days": 62.0,
        "dio_days": 9.0,
        "shares_diluted": 15000,
        "dividend_per_share": 0.0,
    },
    converged=True,
    plug_used=False,
)

SAMPLE_REPORT = VerificationReport(
    passed=True, critical_failures=[], warnings=[], notes=[], period_checks={}
)


def test_writer_creates_file():
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "model.xlsx")
        writer = ExcelWriter(SAMPLE_OUTPUT, SAMPLE_REPORT, "Test Corp", out_path)
        writer.write()
        assert os.path.exists(out_path)
        assert os.path.getsize(out_path) > 5000


def test_writer_creates_required_tabs():
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "model.xlsx")
        writer = ExcelWriter(SAMPLE_OUTPUT, SAMPLE_REPORT, "Test Corp", out_path)
        writer.write()
        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        assert "IS" in wb.sheetnames
        assert "BS" in wb.sheetnames
        assert "CF" in wb.sheetnames
        assert "Assumptions" in wb.sheetnames
        assert "Schedules" in wb.sheetnames
        assert "Sources" in wb.sheetnames
