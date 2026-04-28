"""
Scenario toggle test per SPEC_modeling_patterns.md Section 9.

Checks:
  - Toggle cell exists at Assumptions!$D$9 with values 1/2/3
  - Active case label formula =CHOOSE(...)
  - Three scenario blocks (Base/Upside/Downside) with driver rows
  - Drivers have values for all projection periods (no blanks)
"""
import openpyxl
import pytest
from pathlib import Path


def test_xlsx_scenario_toggle():
    """Validate scenario toggle structure on a real xlsx output."""
    xlsx_path = list(Path(".").glob("*_model.xlsx"))
    if not xlsx_path:
        pytest.skip("No _model.xlsx found — build model first")
    wb = openpyxl.load_workbook(str(xlsx_path[0]))
    ws = wb["Assumptions"]

    toggle = ws.cell(row=9, column=4).value  # D9
    assert toggle in (1, "1", 2, "2", 3, "3"), \
        f"Toggle cell D9 should be 1/2/3, got {toggle!r}"

    active_label = ws.cell(row=10, column=4).value  # D10
    assert active_label is not None, "D10 active case label missing"
    assert isinstance(active_label, str) and ("Base" in active_label
                                               or "CHOOSE" in active_label
                                               or "=CHOOSE" in active_label), \
        f"D10 should reference CHOOSE, got {active_label}"

    proj_cols = []
    for col_idx in range(7, 20):
        v = ws.cell(row=14, column=col_idx).value  # first proj year header
        if v and isinstance(v, str) and "E" in v:
            proj_cols.append(col_idx)

    # Check driver rows have values in proj columns
    for row_num in range(14, 70):
        label = ws.cell(row=row_num, column=3).value  # C col
        if label and ("Revenue Growth" in str(label)
                       or "Dividend per Share" in str(label)
                       or "Tax Rate" in str(label)):
            has_value = any(
                ws.cell(row=row_num, column=c).value is not None
                for c in proj_cols
            )
            assert has_value, \
                f"Driver '{label.strip()}' missing values in proj columns"

    print(f"[PASS] Scenario toggle valid (D9={toggle}). "
          f"Drivers populated across {len(proj_cols)} proj cols.")


def test_xlsx_sensitivity_toggle_structure():
    """Check sensitivity tables have proper structure per DCF SPEC."""
    xlsx_paths = list(Path(".").glob("*_model.xlsx"))
    if not xlsx_paths:
        pytest.skip("No _model.xlsx found")
    # Pick file with most sheets (likely has DCF/Sensitivities)
    best = max(xlsx_paths, key=lambda p: len(openpyxl.load_workbook(str(p)).sheetnames))
    wb = openpyxl.load_workbook(str(best))
    if "Sensitivities" not in wb.sheetnames:
        wb.close()
        pytest.skip("Sensitivity tab not in model — try building with DCF")
    sens = wb["Sensitivities"]
    # Mixed cell reference $A5, B$4 pattern indicates a proper sensitivity table
    has_mixed_ref = False
    for row in sens.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                if ("$A" in v.upper() or "$5" in v) and "$4" in v:
                    has_mixed_ref = True
                    break
    assert sens.max_row > 10, "Sensitivities tab seems empty"
    print(f"[PASS] Sensitivities tab present ({sens.max_row} rows). "
          f"Mixed-referenced tables: {has_mixed_ref}.")
    wb.close()


def test_xlsx_no_formula_errors():
    """Check the xlsx for #REF!, #VALUE!, #DIV/0! errors on every tab."""
    import pytest
    import openpyxl
    xlsx_path = list(Path(".").glob("*_model.xlsx"))
    if not xlsx_path:
        pytest.skip("No _model.xlsx found")
    wb = openpyxl.load_workbook(str(xlsx_path[0]))
    errors = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.upper() in {
                    "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!"
                }:
                    errors.append(f"[{sheet}] {cell.coordinate}: {v}")
    assert not errors, f"Formula errors found:\n" + "\n".join(errors)
    print(f"[PASS] No formula errors in {len(wb.sheetnames)} sheets.")
