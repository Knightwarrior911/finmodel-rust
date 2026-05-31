import json
import openpyxl
from src.source_ledger import SourceLedger
from src.audit_pipeline import annotate_workbook_with_links


def _wb(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WACC"
    ws["A1"] = "Tax Rate"; ws["B1"] = 0.25
    ws["A2"] = "Terminal Growth"; ws["B2"] = 0.025
    ws["A3"] = "Mystery"; ws["B3"] = 0.99
    p = tmp_path / "m.xlsx"; wb.save(p); return p


def test_tiers_colour_cells(tmp_path):
    led = SourceLedger()
    led.record_derived("wacc", "tax rate", None, value=0.25,
                       formula="income_tax/(ni+tax)", inputs=[])
    led.record_assumption("wacc", "terminal growth", None, value=0.025,
                          rationale="GDP proxy", basis="house default")
    cache = {"__ledger__": led.to_json()}
    cp = tmp_path / "cache.json"; cp.write_text(json.dumps(cache), encoding="utf-8")

    xlsx = _wb(tmp_path)
    res = annotate_workbook_with_links(str(xlsx), cache_path=str(cp))
    assert res["derived"] >= 1
    assert res["assumption"] >= 1
    assert res["unverified"] >= 1          # B3 (0.99) is the red catch-all

    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["WACC"]
    assert ws["B1"].comment is not None and "Derived" in ws["B1"].comment.text
    assert ws["B3"].comment is not None and "Unverified" in ws["B3"].comment.text


def test_no_ledger_is_unchanged(tmp_path):
    # When cache has no __ledger__, the new tier keys may be absent or zero and
    # NO catch-all coloring happens. Returns the legacy dict shape.
    xlsx = _wb(tmp_path)
    cp = tmp_path / "cache.json"; cp.write_text(json.dumps({}), encoding="utf-8")
    res = annotate_workbook_with_links(str(xlsx), cache_path=str(cp))
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["WACC"]
    assert ws["B3"].comment is None        # no catch-all when no ledger
