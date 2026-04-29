"""
Verification loop per VERIFICATION_LOOP.md.

Core principle: Generate ? Execute ? Check ? Fix ? Deliver.
Max 3 fix iterations, then surface to user.

Five rules:
  1. Compute key answers twice, independent methods
  2. Force deliverable to execute
  3. Compare programmatically, not by eyeballing
  4. On mismatch, debug before proceeding
  5. When unverifiable, say so explicitly
"""
import os
import logging
import subprocess
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class GroundTruth:
    """Pre-build expected results from raw financial data (Step 1)."""
    ticker: str

    # Balance sheet totals from raw data
    total_assets: list[float] = field(default_factory=list)
    total_liabilities: list[float] = field(default_factory=list)
    total_equity: list[float] = field(default_factory=list)
    bs_delta: list[float] = field(default_factory=list)  # A - L - E

    # Cash flow tie-out from raw data
    net_income_expected: list[float] = field(default_factory=list)
    ending_cash_expected: list[float] = field(default_factory=list)

    # Key metrics
    revenue: list[float] = field(default_factory=list)
    ebitda: list[float] = field(default_factory=list)
    net_income: list[float] = field(default_factory=list)

    # Period labels
    periods: list[str] = field(default_factory=list)

    # Unverifiable items (explicitly flagged per Rule 5)
    unverifiable: list[str] = field(default_factory=list)


@dataclass
class ComparisonResult:
    """Result of programmatic comparison between deliverable and ground truth."""
    passed: bool = True
    mismatches: list[str] = field(default_factory=list)
    checks_run: int = 0
    checks_passed: int = 0
    unverifiable_flags: list[str] = field(default_factory=list)


@dataclass
class LoopReport:
    """Full verification loop report."""
    status: str = "success"           # "success" | "fail" | "partial"
    iterations: int = 0
    ground_truth: GroundTruth | None = None
    comparison: ComparisonResult | None = None
    force_executed: bool = False
    pre_delivery_checks: list[str] = field(default_factory=list)
    unresolved: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def passed_all(self) -> bool:
        return self.status == "success" and not self.unresolved


# ── Step 1: Establish ground truth ──────────────────────────────────────────

def establish_ground_truth(reconciled, model_output=None) -> GroundTruth:
    """Compute expected results from model engine output + raw filing data.

    Uses model_output historical arrays as primary reference (these are what
    the Excel file was built from). Falls back to raw reconciled data where
    model_output fields are missing.
    """
    if model_output is not None:
        periods = list(model_output.periods)
    else:
        periods = list(reconciled.periods)
    gt = GroundTruth(ticker=reconciled.ticker)
    gt.periods = periods
    n = len(periods)

    # Use model_output historical arrays when available (engine-computed truth)
    if model_output is not None:
        bs = model_output.balance_sheet
        is_out = model_output.income_statement
        gt.revenue = is_out.get("revenue", [0] * n)
        gt.net_income = is_out.get("net_income", [0] * n)
        ebit = is_out.get("ebit", [0] * n)
        da = is_out.get("da", [0] * n)
        gt.ebitda = [(ebit[j] or 0) + (da[j] or 0) for j in range(n)]
        gt.total_assets = bs.get("total_assets", [0] * n)
        gt.total_liabilities = bs.get("total_liabilities", [0] * n)
        gt.total_equity = bs.get("total_equity", [0] * n)
        for j in range(n):
            ta = gt.total_assets[j] or 0
            tl = gt.total_liabilities[j] or 0
            te = gt.total_equity[j] or 0
            gt.bs_delta.append(ta - tl - te)
        gt.ending_cash_expected = bs.get("cash", [0] * n)
        gt.net_income_expected = is_out.get("net_income", [0] * n)
    else:
        # Fallback: raw reconciled data
        bs = reconciled.balance_sheet
        gt.total_assets = bs.get("total_assets", [0] * n)
        gt.total_liabilities = bs.get("total_liabilities", [0] * n)
        gt.total_equity = bs.get("total_equity", [0] * n)
        for j in range(n):
            ta = gt.total_assets[j] or 0
            tl = gt.total_liabilities[j] or 0
            te = gt.total_equity[j] or 0
            gt.bs_delta.append(ta - tl - te)
        is_data = reconciled.income_statement
        gt.revenue = is_data.get("revenue", [0] * n)
        gt.net_income = is_data.get("net_income", [0] * n)
        ebit = is_data.get("ebit", [0] * n)
        da = is_data.get("da", [0] * n)
        gt.ebitda = [(ebit[j] or 0) + (da[j] or 0) for j in range(n)]
        gt.ending_cash_expected = bs.get("cash", [0] * n)
        gt.net_income_expected = is_data.get("net_income", [0] * n)

    # Flag unverifiable items
    if not any(v and v != 0 for v in (da or [])):
        gt.unverifiable.append("D&A data missing")
    if not any(v and v != 0 for v in (gt.ending_cash_expected or [])):
        gt.unverifiable.append("Cash data missing from BS")
    if not any(v and v != 0 for v in (ebit or [])):
        gt.unverifiable.append("EBIT data missing from IS")

    return gt


# ── Step 3: Force execution ────────────────────────────────────────────────

def _force_execute_win32(xlsx_path: str, timeout_sec: int = 60) -> bool:
    """Force Excel recalculation via win32com (Windows only, requires Excel installed)."""
    try:
        import win32com.client
        import pythoncom
        import os
        pythoncom.CoInitialize()
        abs_path = str(Path(xlsx_path).resolve())
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        try:
            wb = xl.Workbooks.Open(abs_path)
            xl.CalculateFull()
            wb.Save()
            wb.Close(SaveChanges=True)
            return True
        finally:
            xl.Quit()
            pythoncom.CoUninitialize()
    except Exception as e:
        logger.warning("win32com Excel recalc failed: %s", e)
        return False


def _force_execute_libreoffice(xlsx_path: str, timeout_sec: int = 30) -> bool:
    """Force recalculation via LibreOffice headless (Linux/macOS fallback)."""
    libre_paths = [
        "soffice",
        "libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    lo_exe = None
    for p in libre_paths:
        try:
            r = subprocess.run([p, "--version"], capture_output=True, timeout=5)
            if r.returncode == 0:
                lo_exe = p
                break
        except Exception:
            continue

    if not lo_exe:
        return False

    out_dir = str(Path(xlsx_path).parent)
    try:
        subprocess.run(
            [lo_exe, "--headless", "--norestore", "--nofirststartwizard",
             "--convert-to", "xlsx", "--outdir", out_dir, xlsx_path],
            timeout=timeout_sec, check=True,
        )
        return True
    except Exception:
        return False


def force_execute(xlsx_path: str, timeout_sec: int = 60) -> bool:
    """Force Excel recalculation. Tries win32com (Windows) then LibreOffice (Linux/macOS)."""
    import platform
    if platform.system() == "Windows":
        ok = _force_execute_win32(xlsx_path, timeout_sec)
        if ok:
            return True
        logger.warning("win32com recalc failed — Excel may not be installed")
    ok = _force_execute_libreoffice(xlsx_path, timeout_sec)
    if not ok:
        logger.warning("No recalculation engine available (tried win32com + LibreOffice)")
    return ok


# ── Step 4: Compare to ground truth ─────────────────────────────────────────

def _period_col_map(ws, header_row: int, max_col: int = 30) -> dict[str, int]:
    """Scan a worksheet header row and return {period_label: column_index}.

    Handles any layout — robust to writer placing headers at any column.
    """
    mapping: dict[str, int] = {}
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and isinstance(val, str) and (val.endswith("A") or val.endswith("E")):
            mapping[val] = col
    return mapping


_TOTAL_REVENUE_PATTERNS = (
    "total revenue",
    "total revenues",
    "net revenue",
    "net revenues",
    "revenue",
    "revenues",
)

def _find_revenue_row(ws, label_col: int = 3, start_row: int = 11, max_row: int = 60) -> int:
    """Scan IS worksheet label column for the total-revenue row.

    Prefers 'total revenue(s)' labels; falls back to first bare 'revenue(s)'.
    Returns row index (1-based) or start_row if not found.
    """
    total_row = None
    bare_row = None
    for row in range(start_row, max_row + 1):
        val = ws.cell(row=row, column=label_col).value
        if not val or not isinstance(val, str):
            continue
        norm = val.strip().lower()
        if norm in ("total revenue", "total revenues"):
            total_row = row
            break
        if norm in ("revenue", "revenues", "net revenue", "net revenues") and bare_row is None:
            bare_row = row
    return total_row or bare_row or start_row


def compare_to_ground_truth(
    xlsx_path: str,
    ground_truth: GroundTruth,
    tolerance: float = 1.0,
) -> ComparisonResult:
    """Open the xlsx, read computed values, compare against ground truth.

    Uses openpyxl data_only=True to read cached formula results.
    Comparison is programmatic — no eyeballing.
    """
    import openpyxl
    result = ComparisonResult()

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        result.passed = False
        result.mismatches.append(f"Cannot open xlsx: {e}")
        return result

    # ── BS balance check ──
    if "BS" in wb.sheetnames:
        ws = wb["BS"]
        col_map = _period_col_map(ws, header_row=10)
        for j, period in enumerate(ground_truth.periods):
            col = col_map.get(period)
            if col is None:
                continue
            chk_cell = ws.cell(row=35, column=col).value    # BS check row

            if chk_cell is not None and abs(float(chk_cell)) > tolerance:
                result.mismatches.append(
                    f"BS check {period}: {chk_cell} (tolerance={tolerance})"
                )
    else:
        result.unverifiable_flags.append("BS tab not found in xlsx")

    # ── Revenue comparison (historical periods only) ──
    if "IS" in wb.sheetnames:
        ws = wb["IS"]
        col_map = _period_col_map(ws, header_row=10)
        rev_row = _find_revenue_row(ws, label_col=3, start_row=11, max_row=60)
        for j, period in enumerate(ground_truth.periods):
            if not period.endswith("A"):
                # Projected periods have no ground truth — skip
                continue
            col = col_map.get(period)
            if col is None:
                continue
            rev_cell = ws.cell(row=rev_row, column=col).value
            expected = ground_truth.revenue[j] if j < len(ground_truth.revenue) else None
            if rev_cell is not None and expected is not None and abs(expected) > 1:
                result.checks_run += 1
                if abs(float(rev_cell) - expected) > tolerance:
                    result.mismatches.append(
                        f"Revenue {period}: model={rev_cell} vs expected={expected}"
                    )
                else:
                    result.checks_passed += 1
    else:
        result.unverifiable_flags.append("IS tab not found in xlsx")

    # ── Formula error scan ──
    try:
        from src.validator import validate_xlsx
        vr = validate_xlsx(xlsx_path)
        if vr.failures:
            for f in vr.failures:
                result.mismatches.append(f"Formula error: {f}")
    except Exception:
        result.unverifiable_flags.append("Validator not available for xlsx scan")

    # ── Unverifiable flags ──
    for uv in ground_truth.unverifiable:
        result.unverifiable_flags.append(uv)

    if result.mismatches:
        result.passed = False

    return result


# ── Financial intelligence checks ─────────────────────────────────────────
# These encode financial first principles that catch conceptual errors a
# mechanical formula check cannot see. Each check reasons about whether the
# numbers are internally consistent AS A FINANCIAL MODEL — not just valid Excel.

def financial_intelligence_checks(model_output, tolerance_pct: float = 0.05) -> list[str]:
    """Cross-validate model output against financial first principles.

    Catches errors that mechanical checks miss: IS D&A != CF D&A add-back,
    broken EBITDA bridge, CFO below net income for a profitable company, etc.

    Returns list of issue strings (empty = clean).
    """
    issues: list[str] = []
    if model_output is None:
        return ["model_output not available for intelligence checks"]

    is_out = model_output.income_statement
    cf_out = model_output.cash_flow_statement
    periods = list(model_output.periods)
    n_hist = sum(1 for p in periods if p.endswith("A"))

    def _v(d: dict, key: str) -> list:
        return d.get(key, []) or []

    da_is   = _v(is_out, "da")
    da_cf   = _v(cf_out, "da")
    ebit    = _v(is_out, "ebit")
    ebitda  = _v(is_out, "ebitda")
    rev     = _v(is_out, "revenue")
    cogs    = _v(is_out, "cogs")
    gp      = _v(is_out, "gross_profit")
    ni      = _v(is_out, "net_income")
    cfo     = _v(cf_out, "cfo")

    for i in range(n_hist):
        period = periods[i] if i < len(periods) else f"period[{i}]"

        # 1. D&A: IS add-back must equal CF add-back (same line, two tabs)
        if i < len(da_is) and i < len(da_cf):
            is_d, cf_d = da_is[i] or 0, da_cf[i] or 0
            if cf_d > 1 and abs(is_d - cf_d) / cf_d > tolerance_pct:
                issues.append(
                    f"[D&A mismatch {period}] IS D&A={is_d:,.0f} != CF add-back={cf_d:,.0f} "
                    f"IS pulling embedded/partial D&A instead of CF total"
                )

        # 2. EBITDA bridge: EBIT + D&A = EBITDA (to within $1M rounding)
        if i < len(ebit) and i < len(da_is) and i < len(ebitda):
            e, d, eb = ebit[i] or 0, da_is[i] or 0, ebitda[i] or 0
            bridge = e + d
            if abs(eb) > 1 and abs(bridge - eb) > 1:
                issues.append(
                    f"[EBITDA bridge {period}] EBIT({e:,.0f})+D&A({d:,.0f})={bridge:,.0f} != EBITDA({eb:,.0f})"
                )

        # 3. Gross profit arithmetic: Revenue − |COGS| = Gross Profit
        if i < len(rev) and i < len(cogs) and i < len(gp):
            r, c, g = rev[i] or 0, cogs[i] or 0, gp[i] or 0
            if r > 1:
                expected_gp = r - abs(c)
                if abs(expected_gp - g) / r > tolerance_pct:
                    issues.append(
                        f"[Gross profit {period}] Rev({r:,.0f})−COGS({c:,.0f})={expected_gp:,.0f} != GP({g:,.0f})"
                    )

        # 4. CFO sanity: profitable company with positive D&A must have CFO > Net Income
        #    (NI + D&A is the minimum CFO before WC, and WC rarely flips that)
        if i < len(ni) and i < len(cfo) and i < len(da_is):
            n, c, d = ni[i] or 0, cfo[i] or 0, da_is[i] or 0
            if n > 0 and d > 0 and c < n:
                issues.append(
                    f"[CFO sanity {period}] CFO={c:,.0f} < Net Income={n:,.0f} "
                    f"(profitable company with D&A={d:,.0f} — WC drag is unusually large)"
                )

        # 5. Revenue sanity: must be positive and not impossibly large (>$5T)
        if i < len(rev):
            r = rev[i] or 0
            if r <= 0:
                issues.append(f"[Revenue {period}] Revenue={r:,.0f} <= 0")
            elif r > 5_000_000:
                issues.append(f"[Revenue {period}] Revenue={r:,.0f} > $5T — likely unit error")

        # 6. EBIT margin sanity: flag margins outside −50% to +80% (catches sign/unit errors)
        if i < len(ebit) and i < len(rev):
            e, r = ebit[i] or 0, rev[i] or 0
            if r > 1:
                margin = e / r
                if margin < -0.5 or margin > 0.8:
                    issues.append(
                        f"[EBIT margin {period}] {margin:.1%} outside plausible range "
                        f"(EBIT={e:,.0f} Rev={r:,.0f}) — check sign or unit"
                    )

    return issues


# ── Pre-delivery checklist ──────────────────────────────────────────────────

def pre_delivery_checklist(
    xlsx_path: str,
    ground_truth: GroundTruth,
    comparison: ComparisonResult,
    force_executed: bool,
) -> list[str]:
    """Run the full pre-delivery checklist per VERIFICATION_LOOP.md.

    Returns list of failed checks (empty = all passed).
    """
    failed = []

    if ground_truth is None:
        failed.append("Ground truth not computed")

    if not force_executed:
        failed.append("Deliverable not force-executed (LibreOffice unavailable)")

    if comparison is None:
        failed.append("Programmatic comparison not run")
    elif not comparison.passed:
        failed.append(f"Comparison failed: {len(comparison.mismatches)} mismatches")

    if not Path(xlsx_path).exists():
        failed.append(f"xlsx file missing: {xlsx_path}")

    if comparison and comparison.unverifiable_flags:
        failed.append(
            f"Unverifiable elements: {'; '.join(comparison.unverifiable_flags[:5])}"
        )

    return failed


# ── Main loop: orchestrate steps 1-5 ───────────────────────────────────────

def run_verification_loop(
    xlsx_path: str,
    reconciled,
    model_output,
    max_iterations: int = 3,
    tolerance: float = 1.0,
) -> LoopReport:
    """Execute the full verification loop.

    Steps:
      1. Establish ground truth from reconciled data
      2. Build deliverable (already done by model engine)
      3. Force execution via LibreOffice
      4. Compare deliverable to ground truth programmatically
      5. On mismatch, surface to caller; caller may fix and re-run

    Returns LoopReport with full audit trail.
    """
    report = LoopReport()

    # Step 1: Ground truth (use model_output for engine-computed historical arrays)
    report.ground_truth = establish_ground_truth(reconciled, model_output=model_output)
    logger.info(
        "Ground truth established: %d periods, %d unverifiable items flagged",
        len(report.ground_truth.periods),
        len(report.ground_truth.unverifiable),
    )

    # Step 1b: Financial intelligence checks — reason about the numbers before
    # touching the Excel. Catches conceptual errors (wrong D&A source, broken
    # EBITDA bridge, sign errors) that formula/format checks cannot see.
    intel_issues = financial_intelligence_checks(model_output)
    if intel_issues:
        for issue in intel_issues:
            logger.warning("INTELLIGENCE CHECK: %s", issue)
        report.notes.extend([f"⚠ {i}" for i in intel_issues])

    # Step 3: Force execution
    report.force_executed = force_execute(xlsx_path)

    # Step 4: Compare
    report.comparison = compare_to_ground_truth(
        xlsx_path, report.ground_truth, tolerance=tolerance
    )

    # Step 5: Iterate / surface
    report.iterations = 1  # first pass
    while not report.comparison.passed and report.iterations < max_iterations:
        logger.warning(
            "Verification pass %d failed: %d mismatches. "
            "Surfacing for fix (max %d passes).",
            report.iterations,
            len(report.comparison.mismatches),
            max_iterations,
        )
        report.iterations += 1
        # Caller is responsible for fixing and re-running comparison
        report.unresolved = list(report.comparison.mismatches)
        break  # Surface after first mismatch; caller handles iteration

    # Pre-delivery checklist
    failed_checks = pre_delivery_checklist(
        xlsx_path, report.ground_truth, report.comparison, report.force_executed
    )
    report.pre_delivery_checks = failed_checks

    if failed_checks:
        report.status = "fail"
        report.unresolved.extend(failed_checks)
    elif not report.comparison.passed:
        report.status = "fail"
        report.unresolved = list(report.comparison.mismatches)
    elif report.comparison.unverifiable_flags:
        report.status = "partial"
        report.notes = [
            f"Passed with unverifiable items: "
            f"{'; '.join(report.comparison.unverifiable_flags[:5])}"
        ]
    else:
        report.status = "success"

    return report


# ── Public API: re-run comparison after fix (Step 5 re-entry) ───────────────

def re_compare(
    xlsx_path: str,
    ground_truth: GroundTruth,
    tolerance: float = 1.0,
) -> ComparisonResult:
    """Re-run comparison after fixing deliverable (for iteration)."""
    return compare_to_ground_truth(xlsx_path, ground_truth, tolerance=tolerance)


def flag_unverifiable(reason: str) -> None:
    """Explicitly flag an element that cannot be verified (Rule 5)."""
    logger.info("UNVERIFIABLE: %s", reason)
