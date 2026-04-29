"""Format audit — flag cells where number format doesn't match value type."""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl
import formulas

PATH = 'AAPL_model.xlsx'
xl = formulas.ExcelModel().loads(PATH).finish()
sol = xl.calculate()
recalc = {}
for k, v in sol.items():
    try:
        sheet = k.split(']')[1].split("'")[0]
        cell = k.split('!')[-1]
        try: val = v.value[0][0]
        except: val = v
        recalc[(sheet, cell)] = val
    except Exception:
        pass

wb = openpyxl.load_workbook(PATH, data_only=False)
issues = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            v = recalc.get((sheet_name, cell.coordinate))
            if v is None or not isinstance(v, (int, float)):
                continue
            nf = cell.number_format or ""
            # Heuristic: if format is $/dollar but value is small fraction (<1), likely a percent
            if nf.startswith("$") and abs(v) < 1 and v != 0:
                issues.append(("$-fmt-with-fraction", sheet_name, cell.coordinate, v, nf))
            # If format is % but value is large (>10), likely a $ amount
            if "%" in nf and abs(v) > 10:
                issues.append(("%-fmt-with-large", sheet_name, cell.coordinate, v, nf))
            # If number is negative & shown as dollar, OK
            # Day count rows: format should be #,##0 or General (not $)
            if "Days" in str(ws.cell(row=cell.row, column=3).value or "") and nf.startswith("$"):
                issues.append(("Days-with-$", sheet_name, cell.coordinate, v, nf))

print(f"Format issues found: {len(issues)}")
for typ, sheet, coord, v, nf in issues[:40]:
    label = wb[sheet].cell(row=int(coord[1:] if coord[0].isalpha() else coord), column=3).value if False else ""
    print(f"  [{typ}] [{sheet}] {coord}: val={v} fmt={nf[:30]}")
