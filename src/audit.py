"""Audit AAPL_model.xlsx — compare openpyxl-cached vs formulas-recalc on every cell."""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl
import formulas

PATH = 'AAPL_model.xlsx'

print('Loading + recalcing...')
xl = formulas.ExcelModel().loads(PATH).finish()
sol = xl.calculate()
recalc = {}
for k, v in sol.items():
    # Key format: "'[AAPL_model.xlsx]Sheet'!Cell"
    try:
        sheet = k.split(']')[1].split("'")[0]
        cell = k.split('!')[-1]
        try:
            val = v.value[0][0]
        except (AttributeError, IndexError):
            val = v
        recalc[(sheet, cell)] = val
    except Exception:
        pass

print(f'Recalced {len(recalc)} cells')

wb = openpyxl.load_workbook(PATH, data_only=True)
mismatches = []
checked = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            v_cached = cell.value
            v_recalc = recalc.get((sheet_name, cell.coordinate))
            if v_cached is None and v_recalc is None:
                continue
            if not isinstance(v_cached, (int, float)) or not isinstance(v_recalc, (int, float)):
                continue
            checked += 1
            try:
                if abs(float(v_cached) - float(v_recalc)) > 0.5:
                    mismatches.append((sheet_name, cell.coordinate, v_cached, v_recalc))
            except (ValueError, TypeError):
                pass

print(f'Numeric cells checked: {checked}')
print(f'Mismatches (>$0.5): {len(mismatches)}')
print()
print('First 50 mismatches:')
for m in mismatches[:50]:
    sheet, coord, cached, recalc_v = m
    diff = recalc_v - cached
    print(f'  [{sheet}] {coord}: cached={cached:>15.2f}  recalc={recalc_v:>15.2f}  diff={diff:>+12.2f}')
