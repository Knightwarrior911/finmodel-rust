"""
Output validator (per shared/SPEC_spreadsheet_engineering Section 6).

Checks:
  - Font colors match cell type (blue inputs / black local / green cross-sheet)
  - Two-hop violations (Assumptions! mixed with arithmetic) — FAIL
  - Formula errors (#REF!, #VALUE!, #DIV/0!, #NAME?)  — FAIL
  - Blue cells missing citation comments               — warning
  - Color mismatches                                   — warning

Returns ValidationReport: list of failures + warnings + cell stats.
"""
from dataclasses import dataclass, field
import re


@dataclass
class ValidationReport:
    status: str = "success"            # "success" | "fail"
    failures: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    counts: dict = field(default_factory=dict)
    gridlines_failures: list[str] = field(default_factory=list)


_FORMULA_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

# RGB hex codes for spec font colors (case-insensitive comparisons)
_BLUE  = {"FF0000FF", "0000FF"}
# Brand override per SPEC_excel_formatting Section 14: Ink (#0F1632) used for formulas
# and body text instead of pure #000000. Also includes label gray (#595959) for driver rows.
_BLACK = {"FF000000", "000000", "FF0F1632", "0F1632", "FF595959", "595959"}
_GREEN = {"FF008000", "008000"}
_WHITE = {"FFFFFFFF", "FFFFFF"}   # total/highlight rows — exempt from color rules

# Tabs that are documentation / metadata, not financial calculations
_SKIP_COLOR_SHEETS = {"Sources", "Cover"}


def _normalize_color(c) -> str | None:
    if c is None:
        return None
    val = getattr(c, "rgb", None) or c
    if isinstance(val, str):
        return val.upper().lstrip("#")
    return None


def _is_blue(color: str | None) -> bool:
    return color is not None and color in _BLUE


def _is_black(color: str | None) -> bool:
    return color is not None and color in _BLACK


def _is_green(color: str | None) -> bool:
    return color is not None and color in _GREEN


def _is_two_hop(formula: str) -> bool:
    """Return True if formula mixes Assumptions! cross-ref with arithmetic."""
    if "Assumptions!" not in formula:
        return False
    has_arith = any(op in formula for op in ["*", "+", "-", "/"]) and len(formula) > 1
    if not has_arith:
        return False
    cleaned = formula[1:].strip()
    # Pure lookup: =Assumptions!X1  (no arithmetic after the ref)
    if cleaned.startswith("Assumptions!"):
        after = cleaned[len("Assumptions!"):]
        if not any(op in after for op in ["*", "+", "/"]):
            return False
    # Multiple Assumptions! refs or arithmetic mixed in
    return True


def _extract_refs(formula: str) -> set[str]:
    """Extract cell refs from formula string."""
    refs = set()
    tokens = re.split(r"[+\-*/(),]", formula)
    for token in tokens:
        token = token.replace("=", "").strip()
        m = re.match(r"((\w+!)?\$?[A-Z]{1,3}\$?\d+)", token)
        if m:
            refs.add(m.group(1))
    return refs


def validate_xlsx(path: str) -> ValidationReport:
    """Open xlsx and run spec checks. Returns ValidationReport."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=False)
    rpt = ValidationReport()
    blue = black = green = formulas = errors = 0
    missing_comments = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Gridlines must be off per SPEC_excel_formatting Section 1.1
        sv = getattr(ws, "sheet_view", None)
        if sv is not None:
            show_grid = getattr(sv, "showGridLines", None)
            if show_grid is not False:
                rpt.gridlines_failures.append(
                    f"{sheet}: gridlines visible (should be off)"
                )
        skip_color = sheet in _SKIP_COLOR_SHEETS

        # Build comment set for this sheet
        commented_cells: set[str] = set()
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if getattr(cell, "comment", None) is not None:
                    commented_cells.add(cell.coordinate)

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                font_color = _normalize_color(getattr(cell.font, "color", None))
                is_formula = isinstance(v, str) and v.startswith("=")

                # Label color check (4b): col C text labels must not use data colors
                if (cell.column == 3 and isinstance(v, str) and not is_formula
                        and not isinstance(v, (int, float))):
                    if _is_blue(font_color):
                        rpt.warnings.append(
                            f"[{sheet}] {cell.coordinate}: label text is blue"
                        )
                    elif _is_green(font_color):
                        rpt.warnings.append(
                            f"[{sheet}] {cell.coordinate}: label text is green"
                        )

                # Formula errors → FAIL (even on skipped sheets)
                if isinstance(v, str) and v.upper() in _FORMULA_ERRORS:
                    errors += 1
                    rpt.failures.append(
                        f"[{sheet}] {cell.coordinate}: formula error '{v}'"
                    )
                    continue

                # White font = total/highlight row — exempt from color rules
                if font_color in _WHITE:
                    if is_formula:
                        formulas += 1
                    continue

                # Documentation/metadata sheets — skip color and comment checks
                if skip_color:
                    continue

                if is_formula:
                    formulas += 1
                    has_xref = "!" in v

                    # Two-hop violation → FAIL (blocking per SPEC_spreadsheet_engineering §4)
                    if _is_two_hop(v):
                        rpt.failures.append(
                            f"[{sheet}] {cell.coordinate}: two-hop violation "
                            f"(Assumptions! in arithmetic): {v[:80]}"
                        )

                    # Color expectation
                    if has_xref and not _is_green(font_color):
                        rpt.warnings.append(
                            f"[{sheet}] {cell.coordinate}: cross-sheet formula not green"
                        )
                        green += 1
                    elif has_xref:
                        green += 1
                    elif not _is_black(font_color):
                        rpt.warnings.append(
                            f"[{sheet}] {cell.coordinate}: same-sheet formula not black"
                        )
                        black += 1
                    else:
                        black += 1

                elif isinstance(v, (int, float)):
                    if _is_blue(font_color):
                        blue += 1
                        # Citation comment required on every blue (hardcoded) input
                        if cell.coordinate not in commented_cells:
                            missing_comments += 1
                    else:
                        # Hardcoded number not blue → warning only (may be label/header numeric)
                        rpt.warnings.append(
                            f"[{sheet}] {cell.coordinate}: hardcoded number not blue"
                        )

    # Root-cause error categorization (4c)
    error_cells: set[str] = set()
    for f in rpt.failures:
        parts = f.split("]")
        if len(parts) >= 2:
            coord = parts[1].strip().split(":")[0].strip()
            if coord and " " not in coord:
                error_cells.add(coord)
    all_formulas: dict[str, list[str]] = {}
    if error_cells:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        refs = _extract_refs(cell.value)
                        if refs:
                            coord = f"{sheet}!{cell.coordinate}"
                            all_formulas[coord] = list(refs)
    propagated = 0
    root = 0
    for ec in error_cells:
        ec_ref = ec.split("!")[-1] if "!" in ec else ec
        is_prop = any(
            ec_ref in refs for refs in all_formulas.values()
        )
        if is_prop:
            propagated += 1
        else:
            root += 1
    if propagated:
        rpt.warnings.append(
            f"Propagated errors: {propagated} cell(s) inherit errors. Root cause(s): {root} cell(s)."
        )

    if missing_comments:
        rpt.warnings.append(
            f"Citation comments missing on {missing_comments} blue input cell(s). "
            "Per SPEC_spreadsheet_engineering §5 every hardcoded input requires "
            "a cell comment with cite:{{citationId}}."
        )

    if rpt.gridlines_failures:
        for gf in rpt.gridlines_failures:
            rpt.failures.append(gf)

    rpt.counts = {
        "blue_inputs": blue, "black_formulas": black, "green_xrefs": green,
        "total_formulas": formulas, "formula_errors": errors,
        "missing_comments": missing_comments,
        "warnings": len(rpt.warnings), "failures": len(rpt.failures),
    }
    if rpt.failures:
        rpt.status = "fail"
    return rpt
