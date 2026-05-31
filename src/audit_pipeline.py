"""Audit pipeline — attach CellProvenance to an extraction cache, then post-process
an existing 3-statement xlsx to add a `file#page` source hyperlink on each numeric
cell (click-to-source audit trail, no rendered images).

Public API:
    attach_provenance_to_cache(cache_path, pdf_path) -> int
    annotate_workbook_with_links(xlsx_path, *, cache_path) -> dict
    run_audit(ticker, *, pdf_path=None, xlsx_path=None,
              models_dir=Path("models"),
              cache_dir=Path("extraction_cache")) -> dict

The audit pass is non-invasive: writers and extractors are NOT modified. Provenance
is computed by re-reading the source PDF with PyMuPDF and locating each cached
numeric value via normalize_variants(). Located values link to their exact source
page (#page=N); page-unlocated values with a known source doc link to the doc.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Iterable, Optional

from .provenance import (
    CellProvenance,
    locate_value_in_pdf,
    provenance_dict,
)
from .audit_open import build_uri
from .citations import load_citations


# Statement keys we instrument
_STATEMENT_KEYS = ("income_statement", "balance_sheet", "cash_flow_statement")


def _years(cache: dict[str, Any]) -> list[str]:
    ys = cache.get("years_found") or []
    return [str(y) for y in ys]


def _ticker_from_cache_name(cache_path: Path) -> str:
    name = cache_path.stem  # e.g. "ATCO-B_ST"
    return name.replace("_", "-")


def attach_provenance_to_cache(
    cache_path: str | Path,
    pdf_path: str | Path | None = None,
    *,
    pdf_for_period: dict[str, str] | None = None,
) -> int:
    """Locate every numeric value in the cache inside its source PDF; write
    provenance back into the cache JSON under `__provenance__`.

    Each period (year) is searched in the PDF that actually contains it:
      - pdf_for_period maps "2023" -> path; falls back to pdf_path otherwise.
    A value not found in its period's PDF is stored low_confidence (no bbox)
    rather than risking a wrong-number match.

    Returns the count of values for which a bbox was located.
    """
    import fitz

    cache_path = Path(cache_path)
    cache = json.loads(cache_path.read_text(encoding="utf-8"))
    years = _years(cache)
    pdf_for_period = dict(pdf_for_period or {})
    default_pdf = str(pdf_path) if pdf_path else None

    # Open each distinct PDF once.
    doc_cache: dict[str, Any] = {}

    def _get_doc(path: str | None):
        if not path:
            return None
        if path not in doc_cache:
            try:
                doc_cache[path] = fitz.open(path)
            except Exception:
                doc_cache[path] = None
        return doc_cache[path]

    try:
        provenances: list[CellProvenance] = []
        # Per-(statement, pdf) page hint for speed.
        stmt_hint: dict[tuple[str, str], int] = {}
        for stmt in _STATEMENT_KEYS:
            block = cache.get(stmt) or {}
            for key, values in block.items():
                if not isinstance(values, list):
                    continue
                for idx, val in enumerate(values):
                    if val is None:
                        continue
                    period = years[idx] if idx < len(years) else str(idx)
                    mapped = pdf_for_period.get(period)
                    # Year-mapped mode: search ONLY the period's own report (a
                    # value can't legitimately appear in a different year's
                    # report, and value-only search across the wrong report
                    # yields coincidental false positives). Single-report mode
                    # (no year mapping): use the default PDF for all periods.
                    if pdf_for_period:
                        search_pdf = mapped
                    else:
                        search_pdf = default_pdf
                    doc = _get_doc(search_pdf)
                    page_idx = bbox = raw = None
                    if doc is not None:
                        hk = (stmt, search_pdf)
                        page_idx, bbox, raw = locate_value_in_pdf(
                            doc, val, page_hint=stmt_hint.get(hk),
                        )
                        if page_idx is not None and hk not in stmt_hint:
                            stmt_hint[hk] = page_idx
                    # Citation PDF: where it was found, else the period's mapped
                    # report. Never cite a fallback default for an UNMAPPED
                    # period (that would point at the wrong year's report).
                    if page_idx is not None:
                        cite_pdf = search_pdf
                    else:
                        cite_pdf = mapped or ""
                    provenances.append(CellProvenance(
                        pdf_path=cite_pdf or "",
                        page_index=page_idx if page_idx is not None else 0,
                        bbox=bbox,
                        raw_text=raw or str(val),
                        label=_humanize(key),
                        key=f"{stmt}.{key}",
                        period=period,
                        low_confidence=(page_idx is None),
                    ))
    finally:
        for d in doc_cache.values():
            if d is not None:
                d.close()

    cache["__provenance__"] = provenance_dict(provenances)
    cache_path.write_text(json.dumps(cache, indent=2), encoding="utf-8")

    return sum(1 for p in provenances if not p.low_confidence)


def _humanize(key: str) -> str:
    return key.replace("_", " ").title()


# Cell-value matching for openpyxl post-pass
_NUM_RE = re.compile(r"^-?\d{1,3}(?:[, ]\d{3})*(?:\.\d+)?$")


def _label_tokens(s: str) -> set[str]:
    """Lowercase alpha tokens (len>=3) for fuzzy label matching."""
    return {t for t in re.split(r"[^a-z]+", str(s).lower()) if len(t) >= 3}


def _row_label(ws, cell) -> str:
    """Find the line-item label for a numeric cell: the rightmost non-empty
    string cell to its left on the same row."""
    label = ""
    for c in range(1, cell.column):
        v = ws.cell(row=cell.row, column=c).value
        if isinstance(v, str) and v.strip():
            label = v.strip()
    return label


# Trust-tier font colours (RGB hex) for the ledger-aware Excel render pass.
_TIER_COLOR = {
    "filing": "0000FF", "market": "0000FF", "derived": "595959",
    "assumption": "C55A11", "unverified": "C00000",
}


def _comment_for(tier: str, ref: dict, value: Any) -> Optional[str]:
    """Per-tier cell-comment text. Returns None for tiers (filing/market) whose
    comment is already produced by the existing provenance/citation branches."""
    if tier == "derived":
        return f"Derived: {ref.get('formula','')} = {value}"
    if tier == "assumption":
        return f"Assumption: {ref.get('rationale','')} (basis: {ref.get('basis','')})"
    if tier == "unverified":
        return f"⚠ Unverified: {ref.get('reason','')}"
    return None


def _mark_unverified(cell, openpyxl) -> None:
    """Red catch-all: a numeric cell that matched no ledger/filing/market source.
    Colours the font red and attaches a generic 'no source' comment."""
    old = cell.font
    cell.font = openpyxl.styles.Font(
        name=old.name, size=old.size, bold=old.bold, italic=old.italic,
        color=_TIER_COLOR["unverified"],
    )
    cell.comment = openpyxl.comments.Comment("⚠ Unverified: no source", "ledger")


def build_ledger_index(cache: dict[str, Any]) -> dict[float, list[dict[str, Any]]]:
    """value(float) -> list of {group, field, period, tier, ref, label_tokens}."""
    from .source_ledger import SourceLedger
    led = SourceLedger.from_json(cache.get("__ledger__"))
    idx: dict[float, list[dict[str, Any]]] = {}
    for e in led.entries():
        if e.value is None:
            continue
        idx.setdefault(round(float(e.value), 6), []).append({
            "group": e.group, "field": e.field, "period": e.period,
            "tier": e.tier.value, "ref": e.ref,
            "label_tokens": _label_tokens(e.field.replace("_", " ")),
        })
    return idx


def build_link_indexes(
    cache: dict[str, Any],
) -> tuple[dict[float, list[dict[str, Any]]], dict[float, list[dict[str, Any]]]]:
    """Build the two value->candidates indexes used by every audit post-pass
    (Excel and PPTX): a filing-provenance index and a market-citation index.

    Returns (value_index, market_index):
      value_index[float(value)]      -> filing candidates {full_key, period, pdf,
                                        page_index, label_tokens, low_confidence}
      market_index[round(value, 6)]  -> market candidates {url, label, source,
                                        as_of, label_tokens}
    """
    prov_block = cache.get("__provenance__") or {}
    years = _years(cache)

    value_index: dict[float, list[dict[str, Any]]] = {}
    for full_key, bucket in prov_block.items():
        if not isinstance(bucket, dict):
            continue
        stmt, _, key = full_key.partition(".")
        for period, payload in bucket.items():
            if not isinstance(payload, dict):
                continue
            arr = cache.get(stmt, {}).get(key, [])
            try:
                val = arr[years.index(period)] if period in years else arr[int(period)]
            except Exception:
                val = None
            if val is None:
                continue
            value_index.setdefault(float(val), []).append({
                "full_key": full_key,
                "period": period,
                "pdf": payload.get("pdf_path") or "",
                "page_index": payload.get("page_index"),
                "label_tokens": _label_tokens(payload.get("label", "")),
                "low_confidence": bool(payload.get("low_confidence")),
            })

    market_index: dict[float, list[dict[str, Any]]] = {}
    for mc in load_citations(cache.get("__citations__")):
        market_index.setdefault(round(float(mc.value), 6), []).append({
            "url": mc.url, "label": mc.label, "source": mc.source,
            "as_of": mc.as_of, "label_tokens": _label_tokens(mc.label),
        })
    return value_index, market_index


def annotate_workbook_with_links(
    xlsx_path: str | Path,
    *,
    cache_path: Optional[str | Path] = None,
) -> dict[str, int]:
    """Open an xlsx and, for each numeric cell whose value matches a filing-sourced
    provenance value, attach a `file#page` hyperlink to that number's source.

    Matching is LABEL-AWARE: when a value collides across line items, the cell's
    row label is used to disambiguate (token overlap with the provenance label).

    - Filing value, page known   -> finmodelaudit:page=N&path=<pdf>.
    - Filing value, page unknown -> finmodelaudit page 1.
    - Market-data value (cache __citations__) -> provider https URL.

    The finmodelaudit: scheme is used for filings (not a raw file#page link)
    because Excel drops the #page fragment when the shell opens a local PDF; the
    registered handler (src/audit_open.py) re-launches the browser at the page.
    Market-data links are plain https (Excel opens them in the browser directly).

    Returns {"linked_page", "linked_doc", "linked_market", "total"}. When the
    cache carries a non-empty "__ledger__", a ledger-aware tier pass also runs
    (font colour + tier comment per cell, a red "unverified" catch-all for
    unmatched numeric inputs, and an "Assumptions & Flags" block on a Sources
    sheet) and the return dict additionally carries per-tier counts
    {"filing", "market", "derived", "assumption", "unverified"}. With no ledger
    present this behaviour is fully disabled and the result is unchanged.
    """
    import openpyxl

    if cache_path is None:
        return {"linked_page": 0, "linked_doc": 0, "linked_market": 0, "total": 0}
    cache = json.loads(Path(cache_path).read_text(encoding="utf-8"))
    value_index, market_index = build_link_indexes(cache)

    # Ledger-aware tier rendering is fully gated: when the cache carries no
    # non-empty "__ledger__", none of the new behaviour (tier colouring, red
    # catch-all, Sources summary) runs and the function is byte-identical to its
    # pre-ledger form.
    ledger_present = bool(cache.get("__ledger__", {}).get("entries"))
    ledger_index = build_ledger_index(cache) if ledger_present else {}

    wb = openpyxl.load_workbook(str(xlsx_path))
    linked_page = linked_doc = linked_market = 0
    derived = assumption = filing = market = unverified = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                fv = float(v)
                row_tokens = _label_tokens(_row_label(ws, cell))

                # 0) Ledger tier match (gated) takes priority over filing/market.
                if ledger_present:
                    lcands = ledger_index.get(round(fv, 6))
                    if lcands:
                        lbest = max(
                            lcands,
                            key=lambda c: len(c["label_tokens"] & row_tokens),
                        )
                        # When multiple ledger entries collide on the same value
                        # and none matches the row label, defer to filing/market.
                        if not (len(lcands) > 1 and not (lbest["label_tokens"] & row_tokens)):
                            tier = lbest["tier"]
                            color = _TIER_COLOR.get(tier)
                            if color:
                                old = cell.font
                                cell.font = openpyxl.styles.Font(
                                    name=old.name, size=old.size, bold=old.bold,
                                    italic=old.italic, color=color,
                                )
                            ctext = _comment_for(tier, lbest["ref"], cell.value)
                            if ctext is not None:
                                cell.comment = openpyxl.comments.Comment(ctext, "ledger")
                            if tier == "derived":
                                derived += 1
                            elif tier == "assumption":
                                assumption += 1
                            elif tier == "filing":
                                filing += 1
                            elif tier == "market":
                                market += 1
                            elif tier == "unverified":
                                unverified += 1
                            continue

                # 1) Filing provenance (PDF page) takes priority.
                cands = value_index.get(fv)
                if cands:
                    best = max(cands, key=lambda c: len(c["label_tokens"] & row_tokens))
                    # Require a label match when multiple candidates collide, to
                    # avoid linking a coincidental value (e.g. a computed cell).
                    if len(cands) > 1 and not (best["label_tokens"] & row_tokens):
                        if ledger_present:
                            _mark_unverified(cell, openpyxl)
                            unverified += 1
                        continue
                    if not best["pdf"] or not Path(best["pdf"]).exists():
                        if ledger_present:
                            _mark_unverified(cell, openpyxl)
                            unverified += 1
                        continue
                    page_idx = None if best["low_confidence"] else best["page_index"]
                    page_1based = (page_idx + 1) if page_idx is not None else None
                    cell.hyperlink = build_uri(best["pdf"], page_1based)
                    where = f"page {page_idx + 1}" if page_idx is not None else "source doc"
                    cell.comment = openpyxl.comments.Comment(
                        f"Source: {best['full_key']} {best['period']} ({where})",
                        "audit",
                    )
                    if page_idx is not None:
                        linked_page += 1
                    else:
                        linked_doc += 1
                    filing += 1
                    continue

                # 2) Market-data citation (provider URL, opens in browser).
                mcands = market_index.get(round(fv, 6))
                if not mcands:
                    if ledger_present:
                        _mark_unverified(cell, openpyxl)
                        unverified += 1
                    continue
                mbest = max(mcands, key=lambda c: len(c["label_tokens"] & row_tokens))
                if len(mcands) > 1 and not (mbest["label_tokens"] & row_tokens):
                    if ledger_present:
                        _mark_unverified(cell, openpyxl)
                        unverified += 1
                    continue
                cell.hyperlink = mbest["url"]
                cell.comment = openpyxl.comments.Comment(
                    f"Source: {mbest['label']} via {mbest['source']} "
                    f"(as of {mbest['as_of']})",
                    "audit",
                )
                linked_market += 1
                market += 1

    # Assumptions & Flags summary block on a `Sources` sheet (gated).
    if ledger_present:
        from .source_ledger import SourceLedger, Tier
        ws_src = wb["Sources"] if "Sources" in wb.sheetnames else wb.create_sheet("Sources")
        ws_src.append(["Assumptions & Flags"])
        ws_src.append(["Field", "Tier", "Value", "Rationale / Reason"])
        led = SourceLedger.from_json(cache.get("__ledger__"))
        for e in led.entries_by_tier(Tier.ASSUMPTION, Tier.UNVERIFIED):
            note = e.ref.get("rationale") or e.ref.get("reason") or ""
            ws_src.append([e.field, e.tier.value, e.value, note])

    wb.save(str(xlsx_path))
    return {
        "linked_page": linked_page,
        "linked_doc": linked_doc,
        "linked_market": linked_market,
        "filing": filing,
        "market": market,
        "derived": derived,
        "assumption": assumption,
        "unverified": unverified,
        "total": linked_page + linked_doc + linked_market,
    }


def run_audit(
    ticker: str,
    *,
    pdf_path: Optional[str | Path] = None,
    xlsx_path: Optional[str | Path] = None,
    market_citations: Optional[list] = None,
    models_dir: str | Path = "models",
    cache_dir: str | Path = "extraction_cache",
) -> dict[str, Any]:
    """Full audit pass for a ticker.

    1. Locate cache JSON (extraction_cache/{TICKER}.json with dot->underscore)
    2. Discover source PDFs per period (year in filename) + a default PDF
    3. Attach provenance to cache (each period searched in its own report)
    4. Persist any market-data citations (yfinance/EDGAR) into the cache
    5. If xlsx_path provided, attach source hyperlinks (filing pages + market URLs)
    6. Report honest coverage (located / low_confidence / per-period)
    """
    cache_dir = Path(cache_dir)
    models_dir = Path(models_dir)

    cache_name = ticker.replace(".", "_").replace("-", "_") + ".json"
    cache_path = cache_dir / cache_name
    if not cache_path.exists():
        cache_path = cache_dir / (ticker.replace(".", "_") + ".json")
    if not cache_path.exists():
        return {"ok": False, "error": f"cache not found for {ticker}"}

    cache = json.loads(cache_path.read_text(encoding="utf-8"))
    years = _years(cache)

    # Default PDF: explicit arg, else first ticker-matching PDF in cache_dir
    if pdf_path is None:
        stems = ticker.replace(".", "_").replace("-", "_").split("_")
        for cand in cache_dir.glob("*.pdf"):
            if any(s.lower() in cand.name.lower() for s in stems if len(s) > 2):
                pdf_path = cand
                break

    # Per-period PDF: map each year to a PDF whose filename contains that year.
    # The 2023 report carries FY2023 (+FY2022); the 2025 report carries FY2025
    # (+FY2024) — so a value is only findable in the report that reports it.
    pdfs = list(cache_dir.glob("*.pdf"))
    pdf_for_period: dict[str, str] = {}
    for yr in years:
        # Primary: a report whose filename names this exact year.
        match = next((c for c in pdfs if yr in c.name), None)
        # Comparative fallback: the next year's report carries this year as its
        # prior-period comparative column (e.g. the 2025 report shows FY2024).
        if match is None and yr.isdigit():
            nxt = str(int(yr) + 1)
            match = next((c for c in pdfs if nxt in c.name), None)
        if match is not None:
            pdf_for_period[yr] = str(match)

    if pdf_path is None and not pdf_for_period:
        return {"ok": False, "error": f"PDF not found for {ticker}",
                "cache_path": str(cache_path)}

    n_located = attach_provenance_to_cache(
        cache_path, pdf_path, pdf_for_period=pdf_for_period,
    )

    if market_citations:
        from .citations import persist_citations
        persist_citations(cache_path, market_citations)

    # Coverage stats
    cache = json.loads(cache_path.read_text(encoding="utf-8"))
    prov = cache.get("__provenance__") or {}
    total = low = 0
    per_period_located: dict[str, int] = {}
    # In year-mapped mode, any period without its own report can't be linked.
    if pdf_for_period:
        missing_period_pdfs = [y for y in years if y not in pdf_for_period]
    else:
        missing_period_pdfs = [] if pdf_path else list(years)
    for _k, bucket in prov.items():
        for period, payload in bucket.items():
            total += 1
            if payload.get("low_confidence"):
                low += 1
            else:
                per_period_located[period] = per_period_located.get(period, 0) + 1

    result: dict[str, Any] = {
        "ok": True,
        "ticker": ticker,
        "cache_path": str(cache_path),
        "pdf_path": str(pdf_path) if pdf_path else None,
        "pdf_for_period": pdf_for_period,
        "values_total": total,
        "values_located": n_located,
        "values_low_confidence": low,
        "coverage_pct": round(100.0 * n_located / total, 1) if total else 0.0,
        "located_by_period": per_period_located,
        "missing_period_pdfs": missing_period_pdfs,
    }

    if xlsx_path:
        result["annotated"] = annotate_workbook_with_links(
            xlsx_path, cache_path=cache_path,
        )
        result["xlsx_path"] = str(xlsx_path)

    return result
